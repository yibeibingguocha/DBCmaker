import re
import os
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import traceback
import sys
 
threshold = 3000
stop_exception_handling = False
 
M_LSB = 1
M_MSB = 2
INTEL = 3

title = """
VERSION ""
NS_ : 
    NS_DESC_
    CM_
    BA_DEF_
    BA_
    VAL_
    CAT_DEF_
    CAT_
    FILTER
    BA_DEF_DEF_
    EV_DATA_
    ENVVAR_DATA_
    SGTYPE_
    SGTYPE_VAL_
    BA_DEF_SGTYPE_
    BA_SGTYPE_
    SIG_TYPE_REF_
    VAL_TABLE_
    SIG_GROUP_
    SIG_VALTYPE_
    SIGTYPE_VALTYPE_
    BO_TX_BU_
    BA_DEF_REL_
    BA_REL_
    BA_DEF_DEF_REL_
    BU_SG_REL_
    BU_EV_REL_
    BU_BO_REL_
    SG_MUL_VAL_
BS_:
"""
 
array = [7, 6, 5, 4, 3, 2, 1, 0,
         15, 14, 13, 12, 11, 10, 9, 8,
         23, 22, 21, 20, 19, 18, 17, 16,
         31, 30, 29, 28, 27, 26, 25, 24,
         39, 38, 37, 36, 35, 34, 33, 32,
         47, 46, 45, 44, 43, 42, 41, 40,
         55, 54, 53, 52, 51, 50, 49, 48,
         63, 62, 61, 60, 59, 58, 57, 56,
         71, 70, 69, 68, 67, 66, 65, 64,
         79, 78, 77, 76, 75, 74, 73, 72,
         87, 86, 85, 84, 83, 82, 81, 80,
         95, 94, 93, 92, 91, 90, 89, 88,
         103, 102, 101, 100, 99, 98, 97, 96,
         111, 110, 109, 108, 107, 106, 105, 104,
         119, 118, 117, 116, 115, 114, 113, 112,
         127, 126, 125, 124, 123, 122, 121, 120,
         135, 134, 133, 132, 131, 130, 129, 128,
         143, 142, 141, 140, 139, 138, 137, 136,
         151, 150, 149, 148, 147, 146, 145, 144,
         159, 158, 157, 156, 155, 154, 153, 152,
         167, 166, 165, 164, 163, 162, 161, 160,
         175, 174, 173, 172, 171, 170, 169, 168,
         183, 182, 181, 180, 179, 178, 177, 176,
         191, 190, 189, 188, 187, 186, 185, 184,
         199, 198, 197, 196, 195, 194, 193, 192,
         207, 206, 205, 204, 203, 202, 201, 200,
         215, 214, 213, 212, 211, 210, 209, 208,
         223, 222, 221, 220, 219, 218, 217, 216,
         231, 230, 229, 228, 227, 226, 225, 224,
         239, 238, 237, 236, 235, 234, 233, 232,
         247, 246, 245, 244, 243, 242, 241, 240,
         255, 254, 253, 252, 251, 250, 249, 248,
         263, 262, 261, 260, 259, 258, 257, 256,
         271, 270, 269, 268, 267, 266, 265, 264,
         279, 278, 277, 276, 275, 274, 273, 272,
         287, 286, 285, 284, 283, 282, 281, 280,
         295, 294, 293, 292, 291, 290, 289, 288,
         303, 302, 301, 300, 299, 298, 297, 296,
         311, 310, 309, 308, 307, 306, 305, 304,
         319, 318, 317, 316, 315, 314, 313, 312,
         327, 326, 325, 324, 323, 322, 321, 320,
         335, 334, 333, 332, 331, 330, 329, 328,
         343, 342, 341, 340, 339, 338, 337, 336,
         351, 350, 349, 348, 347, 346, 345, 344,
         359, 358, 357, 356, 355, 354, 353, 352,
         367, 366, 365, 364, 363, 362, 361, 360,
         375, 374, 373, 372, 371, 370, 369, 368,
         383, 382, 381, 380, 379, 378, 377, 376,
         391, 390, 389, 388, 387, 386, 385, 384,
         399, 398, 397, 396, 395, 394, 393, 392,
         407, 406, 405, 404, 403, 402, 401, 400,
         415, 414, 413, 412, 411, 410, 409, 408,
         423, 422, 421, 420, 419, 418, 417, 416,
         431, 430, 429, 428, 427, 426, 425, 424,
         439, 438, 437, 436, 435, 434, 433, 432,
         447, 446, 445, 444, 443, 442, 441, 440,
         455, 454, 453, 452, 451, 450, 449, 448,
         463, 462, 461, 460, 459, 458, 457, 456,
         471, 470, 469, 468, 467, 466, 465, 464,
         479, 478, 477, 476, 475, 474, 473, 472,
         487, 486, 485, 484, 483, 482, 481, 480,
         495, 494, 493, 492, 491, 490, 489, 488,
         503, 502, 501, 500, 499, 498, 497, 496,
         511, 510, 509, 508, 507, 506, 505, 504
         ]
 

# ================== 输出日志 ================== #
 
def print_to_textboxs(*args, end="\n"):
    text = " ".join(str(arg) for arg in args)
    log_text.insert(tk.END, text + end)
    root.update()
    log_text.see(tk.END)
 
 
def print_to_textbox(*args):
    text = " ".join(str(arg) for arg in args)
    log_text.insert(tk.END, text + "\n")
    root.update()
    log_text.see(tk.END)
 

# ================== 表格操作模块 ================== #

def count_worksheets(filename):
    workbook = load_workbook(filename)
    worksheet_count = len(workbook.sheetnames)
    return worksheet_count
 

def get_worksheets(filename):
    workbook = load_workbook(filename)
    worksheets = workbook.sheetnames
    return worksheets
 

def row_to_index(row):
    index = 0
    if isinstance(row, int):
        return row - 1
    if row.isdigit():
        return int(row) - 1
    for char in row:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1
 
 
def insert_signal(arr, signal_name):
    for i in range(len(arr)):
        if signal_name in arr[i]:
            arr.insert(i + 1, 'yes')
            i += 1
    return arr
 

# ================== 文件操作模块 ================== #

def clear_file(filename):
    print_to_textbox("清除dbc文件数据")
    with open(filename, 'w') as file:
        file.truncate()
 

def get_input_directory():
    input_directory = filedialog.askopenfilename(title="选择文件输入目录")
    input_directory_button_entry.delete(0, tk.END)
    input_directory_button_entry.insert(0, input_directory)
    
    # 选择Excel文件后，自动设置DBC输出目录为Excel文件所在目录
    if input_directory and os.path.exists(input_directory):
        # 获取Excel文件所在目录
        excel_dir = os.path.dirname(input_directory)
        output_directory_button_entry.delete(0, tk.END)
        output_directory_button_entry.insert(0, excel_dir + "/output2.dbc")
        
        # 自动检测列名
        auto_detect_and_fill_silent(input_directory)
 

def get_output_directory():
    # 获取当前Excel文件所在目录作为默认打开位置
    excel_path = input_directory_button_entry.get()
    initial_dir = ""
    initial_file = "output2.dbc"
    
    if excel_path and os.path.exists(excel_path):
        initial_dir = os.path.dirname(excel_path)
    
    # 使用asksaveasfilename让用户可以选择目录并编辑文件名
    output_filepath = filedialog.asksaveasfilename(
        title="选择.DBC文件输出路径",
        initialdir=initial_dir,
        initialfile=initial_file,
        defaultextension=".dbc",
        filetypes=[("DBC文件", "*.dbc"), ("所有文件", "*.*")]
    )
    
    if output_filepath:  # 只有用户选择了文件才更新
        output_directory_button_entry.delete(0, tk.END)
        output_directory_button_entry.insert(0, output_filepath)

# ================== 智能列名匹配模块 ================== #

def get_excel_headers(filepath):
    """
    获取Excel文件的列名列表
    """
    try:
        wb = load_workbook(filepath, read_only=True)
        sheet = wb.active
        
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        
        wb.close()
        return headers
    except Exception as e:
        return []


def column_name_to_index(column_name, headers):
    """
    将列名转换为列索引
    如果输入是列号或列字母，直接转换
    如果输入是列名，在headers中查找
    """
    if not column_name:
        return None
    
    column_name = str(column_name).strip()
    
    # 如果是数字，直接返回
    if column_name.isdigit():
        return int(column_name) - 1
    
    # 如果是纯英文列字母（A-Z, AA-ZZ等），且长度不超过3
    if column_name.isalpha() and len(column_name) <= 3 and all(c.isascii() for c in column_name):
        index = 0
        for char in column_name.upper():
            index = index * 26 + (ord(char) - ord('A') + 1)
        return index - 1
    
    # 如果是列名，在headers中查找（支持精确匹配和模糊匹配）
    if headers:
        # 先尝试精确匹配（忽略大小写）
        for idx, header in enumerate(headers):
            if header.lower() == column_name.lower():
                return idx
        
        # 再尝试包含匹配
        for idx, header in enumerate(headers):
            if column_name.lower() in header.lower() or header.lower() in column_name.lower():
                return idx
    
    return None


def auto_detect_columns_silent(filepath):
    """
    静默自动检测Excel文件中的列名并智能匹配（不输出日志）
    返回匹配结果字典，包含列名和列索引
    """
    try:
        wb = load_workbook(filepath, read_only=True)
        sheet = wb.active
        
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        
        wb.close()
        
        if not headers:
            return None
        
        match_rules = {
            'message_name': ['报文名称', '报文名', '消息名称', '消息名', 'message name', 'message_name', 'msg name', '帧名称', '帧名'],
            'message_id': ['报文id', '消息id', 'message id', 'message_id', 'can id', 'can_id', 'id', '帧id', '帧 ID'],
            'signal_name': ['信号名称', '信号名', 'signal name', 'signal_name', 'sig name', '参数名称', '参数名'],
            'start_bit': ['起始位', '起始bit', 'start bit', 'start_bit', '开始位', '起始位置'],
            'length': ['长度', '信号长度', 'length', 'bit length', '位长', '位数'],
            'precision': ['精度', '因子', 'precision', 'factor', '分辨率', 'scale'],
            'offset': ['偏移', '偏移量', 'offset', '初值', '初始值'],
            'comment': ['description', '描述']
        }
        
        matched_columns = {
            'message_name': None,
            'message_id': None,
            'signal_name': None,
            'start_bit': None,
            'length': None,
            'precision': None,
            'offset': None,
            'comment': None
        }
        
        for field, keywords in match_rules.items():
            best_match = None
            best_score = 0
            
            for col_idx, header in enumerate(headers):
                if not header:
                    continue
                
                header_lower = header.lower()
                score = 0
                
                for keyword in keywords:
                    if header_lower == keyword.lower():
                        score = 100
                        break
                    elif keyword.lower() in header_lower:
                        score = max(score, 50)
                    elif header_lower in keyword.lower():
                        score = max(score, 30)
                
                if score > best_score:
                    best_score = score
                    best_match = header
            
            if best_score >= 30:
                matched_columns[field] = best_match
        
        return matched_columns
        
    except Exception as e:
        return None


def apply_matched_columns_silent(matched_columns):
    """
    将自动匹配的结果应用到输入框中（使用列名而非列号）
    """
    if matched_columns is None:
        return False
    
    fields_mapping = {
        'message_name': input_message_name_entry,
        'message_id': input_message_id_entry,
        'signal_name': input_signal_name_entry,
        'start_bit': input_start_bit_entry,
        'length': input_length_entry,
        'precision': input_precision_entry,
        'offset': input_offset_entry,
        'comment': input_comment_entry
    }
    
    all_filled = True
    for field, entry_widget in fields_mapping.items():
        col_name = matched_columns.get(field)
        if col_name:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, col_name)
        else:
            if field != 'comment':
                all_filled = False
    
    return all_filled


def auto_detect_and_fill_silent(filepath):
    """
    静默自动检测并填充（不显示任何日志和弹窗）
    """
    matched_columns = auto_detect_columns_silent(filepath)
    
    if matched_columns:
        apply_matched_columns_silent(matched_columns)

# ================== 函数操作模块 ================== #

def on_closing():
    if messagebox.askokcancel("退出程序", "确定要退出吗?"):
        root.destroy()
        sys.exit()
 

def set_lsb_key():
    global storage_mode
    storage_mode = M_LSB
    print_to_textbox("信号存储方式为Motorola最低有效字节（LSB）")

def set_msb_key():
    global storage_mode
    storage_mode = M_MSB
    print_to_textbox("信号存储方式为Motorola最高有效字节（MSB）")

def set_intel_key():
    global storage_mode
    storage_mode = INTEL
    print_to_textbox("信号存储方式为INTEL最高有效字节（INTEL）")
 

def has_chinese_or_empty(text):
    if not text:
        return True
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    result = re.search(pattern, text)
    if result:
        return True
    else:
        return False
 

def sort_sg_lines(sg_lines):
    pattern = r':\s*(\d+)'
    key_func = lambda line: int(re.findall(pattern, line)[0]) if re.findall(pattern, line) else 0
    sorted_sg_lines = sorted(sg_lines, key=key_func)
    return sorted_sg_lines
 

def sort_dbc_lines(input_list, output_list):
    input_list = [line.strip() for line in input_list]
    bo_lines = [line for line in input_list if line.startswith('BO_')]
    sg_lines = [line for line in input_list if line.startswith('SG_')]
    bo_sg_dict = {}
    current_bo = None
    for line in input_list:
        if line.startswith('BO_'):
            current_bo = line
            bo_sg_dict[current_bo] = []
        elif line.startswith('SG_'):
            bo_sg_dict[current_bo].append(line)
    for bo, sg_lines in bo_sg_dict.items():
        sorted_sg_lines = sort_sg_lines(sg_lines)
        bo_sg_dict[bo] = sorted_sg_lines
    for bo in bo_lines:
        output_list.append('\n' + bo)
        output_list.extend([' ' + sg_line for sg_line in bo_sg_dict[bo]])
 

def convert_to_dbc(worksheet_list, wb):
    global storage_mode
    dbc_rows = []
    comment_rows = []
    bu_nodes = set()  # 收集所有节点名称
    worksheet_names = wb.sheetnames
    worksheet_count = len(worksheet_names)
    has_comment_column = worksheet_list[7] is not None
 
    print_to_textbox("活动表格数量：", worksheet_count, " ,活动表格名数组：", worksheet_names)
    if has_comment_column:
        print_to_textbox("✓ 已启用信号注释功能")
    
    for index in range(worksheet_count):
        current_id_error = []
        repetition_id_array = []
        repetition_id = 0
        previous_row = None
        empty_row_count = 0  # 每个sheet开始时重置空行计数器
        
        # 使用sheet名作为节点名（去除空格和特殊字符）
        sheet_name = worksheet_names[index]
        node_name = re.sub(r'[^a-zA-Z0-9_]', '_', sheet_name.replace(' ', '_'))
        # 确保节点名不以数字开头
        if node_name and node_name[0].isdigit():
            node_name = 'N_' + node_name
        if not node_name:
            node_name = 'Node_' + str(index)
        
        bu_nodes.add(node_name)
 
        sheet = wb[worksheet_names[index]]
        row_count = sheet.max_row
        print_to_textbox("\n========== 处理Sheet:", worksheet_names[index], "(节点:", node_name, ")==========")
        print_to_textbox("表格中的行数为:", row_count)
        for row in sheet.iter_rows(min_row=2, max_row=threshold, values_only=True):
 
            if all(cell is None for cell in row):
                empty_row_count += 1
            else:
                empty_row_count = 0
 
            if empty_row_count >= 5:
                print_to_textbox(f"检测到连续{empty_row_count}个空行，停止处理当前Sheet")
                break
 
            current_row = row[worksheet_list[1]]
            if (row[worksheet_list[0]] is None and row[worksheet_list[2]] is None) or row[worksheet_list[1]] is None:
                continue
 
            current_row = str(current_row).replace(" ", "")
            can_id_int = int(current_row, 16)
            
            if current_row != previous_row:
                print_to_textboxs(current_row, end=", ")
                
                ecu_str = node_name  # 使用当前sheet对应的节点名
                for num_cell, cell in enumerate(row):
                    if cell is not None and "TX" in str(cell).upper():
                        header_value = sheet.cell(row=1, column=num_cell + 1).value
                        if header_value and "TBOX" in str(header_value):
                            ecu_str = "TBOX"
                            bu_nodes.add("TBOX")
                            break
                        else:
                            continue
                    else:
                        ecu_str = node_name
                if row[worksheet_list[0]] is None:
                    continue
                msg_name = str(row[worksheet_list[0]])
                
                dbc_rows.append(
                    f"\nBO_ {can_id_int} {msg_name}: 8 {ecu_str}")
 
            previous_row = current_row
 
            signal_name_raw = row[worksheet_list[2]]
            if signal_name_raw is None or has_chinese_or_empty(str(signal_name_raw)) or 'Reserved' in str(signal_name_raw):
                continue
 
            signal_name = str(signal_name_raw).replace('\n', '')
            signal_name = signal_name.replace(' ', '')
            signal_name = re.sub(r'[^a-zA-Z0-9]', '_', signal_name)
 
            start_bitvalue = str(row[worksheet_list[3]])
            try:
                start_bit = int(start_bitvalue)
            except ValueError:
                current_id_error.append(current_row)
                continue
 
            lengthvalue = str(row[worksheet_list[4]])
            try:
                length = int(lengthvalue)
            except ValueError:
                current_id_error.append(current_row)
                continue
 
            if storage_mode == M_LSB:
                end_bit = array[array.index(start_bit) - length + 1]
            elif storage_mode == M_MSB:
                end_bit = array[array.index(start_bit) + length - 1]
            else:
                end_bit = start_bit + length - 1
 
            precision = str(row[worksheet_list[5]])
            if precision == 'None':
                precision = '1'
            offset = str(row[worksheet_list[6]])
            if offset == 'None':
                offset = '0'
 
            # 处理注释
            comment_text = ""
            if has_comment_column and worksheet_list[7] is not None and row[worksheet_list[7]] is not None:
                comment_text = str(row[worksheet_list[7]]).strip()
                # 清理注释中的特殊字符，避免DBC格式错误
                comment_text = comment_text.replace('"', "'").replace('\n', ' ').replace('\r', ' ')
 
            # 查找当前CAN ID对应的BO_行并插入信号
            target_bo_found = False
            for i, row_data in enumerate(dbc_rows):
                if row_data.startswith('BO_') and f"BO_ {can_id_int} " in row_data:
                    # 在这个BO_后面插入信号
                    dbc_rows.insert(i + 1, f" SG_ {signal_name} : {end_bit}|{length}@0+ ({precision},{offset}) [0|0] \"\" Vector__XXX")
                    target_bo_found = True
                    # 添加注释
                    if comment_text:
                        comment_rows.append(f'CM_ SG_ {can_id_int} {signal_name} "{comment_text}";')
                    break
            
            # 如果没有找到对应的BO_，添加到末尾
            if not target_bo_found:
                dbc_rows.append(f" SG_ {signal_name} : {end_bit}|{length}@0+ ({precision},{offset}) [0|0] \"\" Vector__XXX")
                if comment_text:
                    comment_rows.append(f'CM_ SG_ {can_id_int} {signal_name} "{comment_text}";')
 
        print_to_textbox(f'\n{worksheet_names[index]} 转换完成!-----------转换进度{str((index + 1) * 100 / worksheet_count)}%')
        print_to_textbox(f"重复报文 ID : {repetition_id_array}")
        print_to_textbox(f"无法转换 ID : {current_id_error}")
 
    # 生成BU_行，列出所有节点
    bu_line = "BU_: " + " ".join(sorted(bu_nodes))
    
    dbc_rows_output = []
    sort_dbc_lines(dbc_rows, dbc_rows_output)
    print_to_textbox("\n 完成信号排版 -------")
    print_to_textbox(f"共识别 {len(bu_nodes)} 个节点: {', '.join(sorted(bu_nodes))}")
    
    # worksheet_list[10] 是输出DBC文件的完整路径
    output_filepath = worksheet_list[10]
    
    # 使用GBK编码保存文件，支持21003个汉字，被CAN分析工具广泛支持
    # errors='replace' 作为保护机制，将无法编码的字符替换为?
    with open(output_filepath, 'w', encoding='gbk', newline='', errors='replace') as f:
        f.write(title)
        f.write(bu_line + "\n\n")  # 写入节点列表
        f.write('\n'.join(dbc_rows_output) + '\n')
        # 写入注释（在所有信号定义之后）
        if comment_rows:
            f.write('\n'.join(comment_rows) + '\n')
            print_to_textbox(f"✓ 已写入 {len(comment_rows)} 条信号注释")
 

def submit_main():
    input_data0 = input_message_name_entry.get().strip()
    input_data1 = input_message_id_entry.get().strip()
    input_data2 = input_signal_name_entry.get().strip()
    input_data3 = input_start_bit_entry.get().strip()
    input_data4 = input_length_entry.get().strip()
    input_data5 = input_precision_entry.get().strip()
    input_data6 = input_offset_entry.get().strip()
    input_data7 = input_comment_entry.get().strip() if input_comment_entry.get().strip() else None
    input_directory = input_directory_button_entry.get()
    output_filepath = output_directory_button_entry.get()
    
    # 如果输出路径为空，使用默认路径
    if not output_filepath:
        excel_dir = os.path.dirname(input_directory) if input_directory else ""
        output_filepath = excel_dir + "/output2.dbc" if excel_dir else "output2.dbc"
 
    worksheet_list = [input_data0, input_data1, input_data2, input_data3, input_data4, input_data5, input_data6, input_data7,
                      input_directory, output_filepath]
 
    # 获取Excel列头用于列名转换
    headers = get_excel_headers(input_directory)
    
    if not headers:
        print_to_textbox("错误: 无法读取Excel文件列头，请检查文件是否正确")
        return
    
    print_to_textbox(f"Excel列头: {headers}")
    print_to_textbox(f"总列数: {len(headers)}")
    
    # 将列名转换为列索引（只转换前8个字段）
    converted_list = []
    field_names = ['报文名称', '报文ID', '信号名称', '起始地址', '信号长度', '精度', '偏移量', '注释']
    
    for idx, item in enumerate(worksheet_list[:8]):
        if item is None or item == '':
            converted_list.append(None)
            continue
            
        col_index = column_name_to_index(item, headers)
        if col_index is None:
            print_to_textbox(f"❌ 错误: 无法识别'{field_names[idx]}'的列 '{item}'")
            print_to_textbox(f"   请检查输入是否正确，可用列: {headers}")
            return
        
        if col_index >= len(headers):
            print_to_textbox(f"❌ 错误: '{field_names[idx]}'的列索引 {col_index} 超出范围 (最大索引: {len(headers)-1})")
            print_to_textbox(f"   输入的列: {item}")
            return
        
        converted_list.append(col_index)
        print_to_textbox(f"✓ {field_names[idx]}: '{item}' -> 列索引 {col_index} (列名: {headers[col_index]})")
    
    # 添加文件路径（保持原样，不转换）
    converted_list.append(worksheet_list[8])  # 输入Excel文件路径
    converted_list.append(worksheet_list[9])  # 输出目录路径
    worksheet_list = converted_list
 
    print_to_textbox("\n========== 列映射确认 ==========")
    print_to_textbox("输入报文名所在列:", input_data0, "-> 索引:", worksheet_list[0])
    print_to_textbox("输入报文ID所在列:", input_data1, "-> 索引:", worksheet_list[1])
    print_to_textbox("输入信号名所在列:", input_data2, "-> 索引:", worksheet_list[2])
    print_to_textbox("输入起始地址所在列:", input_data3, "-> 索引:", worksheet_list[3])
    print_to_textbox("输入信号长度所在列:", input_data4, "-> 索引:", worksheet_list[4])
    print_to_textbox("输入精度所在列:", input_data5, "-> 索引:", worksheet_list[5])
    print_to_textbox("输入偏移量所在列:", input_data6, "-> 索引:", worksheet_list[6])
    if worksheet_list[7] is not None:
        print_to_textbox("输入注释所在列:", input_data7, "-> 索引:", worksheet_list[7])
    else:
        print_to_textbox("注释列: 未设置")
    print_to_textbox("================================\n")
    
    print_to_textbox("XLSX文件输入目录:", input_directory)
    print_to_textbox("DBC文件输出路径:", output_filepath)
 
    if storage_mode == M_LSB:
        print_to_textbox("信号存储方式为LSB")
    elif storage_mode == M_MSB:
        print_to_textbox("信号存储方式为MSB")
    else:
        print_to_textbox("信号存储方式为INTEL")
 
    # worksheet_list[8] 是输入Excel文件路径
    wb = openpyxl.load_workbook(worksheet_list[8])
 
    # worksheet_list[9] 是输出DBC文件的完整路径
    filepath = worksheet_list[9]
 
    # 使用GBK编码创建初始文件（支持21003个汉字，兼容性好）
    if os.path.exists(filepath):
        with open(filepath, 'w', encoding='gbk', errors='replace') as dbc_file:
            dbc_file.truncate(0)
    else:
        with open(filepath, 'w', encoding='gbk', errors='replace'):
            pass
 
    # 将输出文件路径添加到worksheet_list供convert_to_dbc使用
    worksheet_list.append(filepath)
 
    try:
        convert_to_dbc(worksheet_list, wb)
        print_to_textbox("-----------XlSX转换DBC完毕!")
    except Exception as e:
        error_traceback = traceback.format_exc()
        print_to_textbox("\n！！！代码转化发生错误！！！ \n: ", error_traceback)
        print_to_textbox("\n")
        print_to_textbox("-----------XlSX转换DBC失败!")

root = tk.Tk()
root.geometry("720x410")
root.protocol("WM_DELETE_WINDOW", on_closing)
 
storage_mode = 1
 
log_text = tk.Text(root, width=72, height=15)
log_text.pack()
log_text.grid(row=0, column=0, columnspan=3, padx=15, pady=5)
print_to_textbox("编译日志 : \r\n默认信号存储方式为最低有效字节（LSB）")
print_to_textbox("提示: 选择Excel文件后将自动识别列名，可直接修改或使用默认值")
print_to_textbox("提示: DBC文件默认名为output2.dbc，可在保存时修改\n")
 
input_message_name_label = tk.Label(root, text="输入报文名称所在列:")
input_message_name_label.place(x=10, y=215)
input_message_name_entry = tk.Entry(root)
input_message_name_entry.place(x=175, y=215)
input_message_name_entry.configure(width=24)
 
input_message_id_label = tk.Label(root, text="输入报文ID所在列:")
input_message_id_label.place(x=360, y=215)
input_message_id_entry = tk.Entry(root)
input_message_id_entry.place(x=525, y=215)
input_message_id_entry.configure(width=24)
 
input_signal_name_label = tk.Label(root, text="输入信号名称所在列:")
input_signal_name_label.place(x=10, y=245)
input_signal_name_entry = tk.Entry(root)
input_signal_name_entry.place(x=175, y=245)
input_signal_name_entry.configure(width=24)
 
input_start_bit_label = tk.Label(root, text="输入起始地址所在列:")
input_start_bit_label.place(x=360, y=245)
input_start_bit_entry = tk.Entry(root)
input_start_bit_entry.place(x=525, y=245)
input_start_bit_entry.configure(width=24)
 
input_length_label = tk.Label(root, text="输入信号长度所在列:")
input_length_label.place(x=10, y=275)
input_length_entry = tk.Entry(root)
input_length_entry.place(x=175, y=275)
input_length_entry.configure(width=24)
 
input_precision_label = tk.Label(root, text="输入精度所在列:")
input_precision_label.place(x=360, y=275)
input_precision_entry = tk.Entry(root)
input_precision_entry.place(x=525, y=275)
input_precision_entry.configure(width=24)
 
input_offset_label = tk.Label(root, text="输入偏移量所在列:")
input_offset_label.place(x=10, y=305)
input_offset_entry = tk.Entry(root)
input_offset_entry.place(x=175, y=305)
input_offset_entry.configure(width=24)
 
input_comment_label = tk.Label(root, text="输入注释所在列(可选):")
input_comment_label.place(x=360, y=305)
input_comment_entry = tk.Entry(root)
input_comment_entry.place(x=525, y=305)
input_comment_entry.configure(width=24)
 
input_directory_button_open = tk.Button(root, text="选择XLSX文件输入目录", command=get_input_directory)
input_directory_button_open.place(x=10, y=340)
input_directory_button_entry = tk.Entry(root)
input_directory_button_entry.place(x=175, y=340)
input_directory_button_entry.configure(width=24)
 
output_directory_button_open = tk.Button(root, text="选择.DBC文件输出路径", command=get_output_directory)
output_directory_button_open.place(x=360, y=340)
output_directory_button_entry = tk.Entry(root)
output_directory_button_entry.place(x=525, y=340)
output_directory_button_entry.configure(width=24)
 
button1 = tk.Button(root, text="LSB", command=set_lsb_key)
button1.place(x=25, y=375)
button1.configure(width=10)
 
button2 = tk.Button(root, text="MSB", command=set_msb_key)
button2.place(x=137, y=375)
button2.configure(width=10)
 
button3 = tk.Button(root, text="INTEL", command=set_intel_key)
button3.place(x=255, y=375)
button3.configure(width=10)
 
submit_button = tk.Button(root, text="生成DBC文件", command=submit_main)
submit_button.place(x=550, y=10)
submit_button.configure(width=20, height=10)
 
root.mainloop()