# -*- coding: utf-8 -*-
"""
DBC转化工具 - 自动化测试脚本
用于验证收发方向功能的正确性
"""

import openpyxl
import os
import sys

def create_test_excel():
    """创建测试用的Excel文件"""
    print("创建测试Excel文件...")
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: ADASCAN - 发送报文
    ws1 = wb.active
    ws1.title = 'ADASCAN'
    
    headers = ['报文名称', '报文ID', '信号名称', '起始位', '长度', '精度', '偏移量', '收发方向']
    ws1.append(headers)
    
    # 添加测试数据 - 发送报文
    ws1.append(['EngineData', '0x100', 'RPM', 0, 16, 1, 0, 'T'])
    ws1.append(['EngineData', '0x100', 'Temperature', 16, 8, 0.1, -40, 'T'])
    ws1.append(['VehicleSpeed', '0x200', 'Speed', 0, 16, 0.01, 0, 'T'])
    
    # Sheet 2: BCAN - 接收报文
    ws2 = wb.create_sheet('BCAN')
    ws2.append(headers)
    
    # 添加测试数据 - 接收报文（相同ID但不同方向）
    ws2.append(['BrakeStatus', '0x100', 'Pressure', 0, 8, 1, 0, 'R'])
    ws2.append(['BrakeStatus', '0x100', 'Warning', 8, 1, 1, 0, 'R'])
    ws2.append(['DoorStatus', '0x300', 'LeftDoor', 0, 1, 1, 0, 'R'])
    ws2.append(['DoorStatus', '0x300', 'RightDoor', 1, 1, 1, 0, 'R'])
    
    # Sheet 3: PT_CAN - 混合方向
    ws3 = wb.create_sheet('PT_CAN')
    ws3.append(headers)
    
    # 添加测试数据 - 混合收发
    ws3.append(['TxMessage', '0x400', 'Data1', 0, 8, 1, 0, 'T'])
    ws3.append(['RxMessage', '0x500', 'Data2', 0, 8, 1, 0, 'R'])
    ws3.append(['TxMessage2', '0x600', 'Data3', 0, 8, 1, 0, 'T'])
    
    filename = 'test_direction_demo.xlsx'
    wb.save(filename)
    print(f"✓ 测试文件已创建: {filename}")
    return filename


def verify_dbc_output(dbc_file):
    """验证生成的DBC文件"""
    print("\n验证DBC文件内容...")
    
    if not os.path.exists(dbc_file):
        print(f"✗ DBC文件不存在: {dbc_file}")
        return False
    
    with open(dbc_file, 'r', encoding='gbk') as f:
        content = f.read()
    
    # 检查关键内容
    checks = [
        ('BU_:', '节点列表'),
        ('ADASCAN', 'ADASCAN节点'),
        ('BCAN', 'BCAN节点'),
        ('PT_CAN', 'PT_CAN节点'),
        ('BO_ 256', 'EngineData报文'),
        ('BO_ 512', 'VehicleSpeed报文'),
        ('BO_ 768', 'BrakeStatus报文'),
        ('接收报文(R)', '接收报文注释'),
    ]
    
    all_passed = True
    for keyword, description in checks:
        if keyword in content:
            print(f"  ✓ {description}: 找到")
        else:
            print(f"  ✗ {description}: 未找到")
            all_passed = False
    
    # 统计报文数量
    bo_count = content.count('BO_ ')
    cm_count = content.count('CM_ BO_')
    sg_count = content.count('SG_ ')
    
    print(f"\n统计信息:")
    print(f"  报文数量 (BO_): {bo_count}")
    print(f"  信号数量 (SG_): {sg_count}")
    print(f"  注释数量 (CM_): {cm_count}")
    
    if all_passed:
        print("\n✓ 所有检查通过!")
    else:
        print("\n✗ 部分检查失败，请检查DBC文件")
    
    return all_passed


def main():
    """主函数"""
    print("=" * 60)
    print("DBC转化工具 - 收发方向功能测试")
    print("=" * 60)
    
    # 步骤1: 创建测试Excel
    excel_file = create_test_excel()
    
    print("\n" + "=" * 60)
    print("测试说明:")
    print("=" * 60)
    print("1. 运行 DBC转化工具v1.0.0.0.py")
    print(f"2. 选择文件: {excel_file}")
    print("3. 确认'收发方向所在列'自动识别为 '收发方向' 或 'H'")
    print("4. 点击'生成DBC文件'")
    print("5. 检查生成的 output2.dbc 文件")
    print("\n预期结果:")
    print("  - ADASCAN的0x100报文标记为发送(T)")
    print("  - BCAN的0x100报文标记为接收(R)，带注释")
    print("  - PT_CAN包含发送和接收两种报文")
    print("=" * 60)
    
    print("\n提示: 请手动运行工具并生成DBC文件后，按回车继续验证...")
    input()
    
    # 步骤2: 验证输出
    dbc_file = 'output2.dbc'
    if os.path.exists(dbc_file):
        verify_dbc_output(dbc_file)
        
        print("\n" + "=" * 60)
        print("查看生成的DBC文件内容:")
        print("=" * 60)
        with open(dbc_file, 'r', encoding='gbk') as f:
            print(f.read())
    else:
        print(f"\n✗ 未找到DBC文件: {dbc_file}")
        print("请先运行工具生成DBC文件")


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n✗ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
