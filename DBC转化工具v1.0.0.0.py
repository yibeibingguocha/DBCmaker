import re
import os
import openpyxl
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import tkinter.ttk as ttk
import traceback
import sys
 
threshold = 3000
stop_exception_handling = False
 
M_LSB = 1
M_MSB = 2
INTEL = 3

title = """
VERSION ""
NS_ : 
    NS_DESC_
    CM_
    BA_DEF_
    BA_
    VAL_
    CAT_DEF_
    CAT_
    FILTER
    BA_DEF_DEF_
    EV_DATA_
    ENVVAR_DATA_
    SGTYPE_
    SGTYPE_VAL_
    BA_DEF_SGTYPE_
    BA_SGTYPE_
    SIG_TYPE_REF_
    VAL_TABLE_
    SIG_GROUP_
    SIG_VALTYPE_
    SIGTYPE_VALTYPE_
    BO_TX_BU_
    BA_DEF_REL_
    BA_REL_
    BA_DEF_DEF_REL_
    BU_SG_REL_
    BU_EV_REL_
    BU_BO_REL_
    SG_MUL_VAL_
BS_:
"""
 
array = [7, 6, 5, 4, 3, 2, 1, 0,
         15, 14, 13, 12, 11, 10, 9, 8,
         23, 22, 21, 20, 19, 18, 17, 16,
         31, 30, 29, 28, 27, 26, 25, 24,
         39, 38, 37, 36, 35, 34, 33, 32,
         47, 46, 45, 44, 43, 42, 41, 40,
         55, 54, 53, 52, 51, 50, 49, 48,
         63, 62, 61, 60, 59, 58, 57, 56,
         71, 70, 69, 68, 67, 66, 65, 64,
         79, 78, 77, 76, 75, 74, 73, 72,
         87, 86, 85, 84, 83, 82, 81, 80,
         95, 94, 93, 92, 91, 90, 89, 88,
         103, 102, 101, 100, 99, 98, 97, 96,
         111, 110, 109, 108, 107, 106, 105, 104,
         119, 118, 117, 116, 115, 114, 113, 112,
         127, 126, 125, 124, 123, 122, 121, 120,
         135, 134, 133, 132, 131, 130, 129, 128,
         143, 142, 141, 140, 139, 138, 137, 136,
         151, 150, 149, 148, 147, 146, 145, 144,
         159, 158, 157, 156, 155, 154, 153, 152,
         167, 166, 165, 164, 163, 162, 161, 160,
         175, 174, 173, 172, 171, 170, 169, 168,
         183, 182, 181, 180, 179, 178, 177, 176,
         191, 190, 189, 188, 187, 186, 185, 184,
         199, 198, 197, 196, 195, 194, 193, 192,
         207, 206, 205, 204, 203, 202, 201, 200,
         215, 214, 213, 212, 211, 210, 209, 208,
         223, 222, 221, 220, 219, 218, 217, 216,
         231, 230, 229, 228, 227, 226, 225, 224,
         239, 238, 237, 236, 235, 234, 233, 232,
         247, 246, 245, 244, 243, 242, 241, 240,
         255, 254, 253, 252, 251, 250, 249, 248,
         263, 262, 261, 260, 259, 258, 257, 256,
         271, 270, 269, 268, 267, 266, 265, 264,
         279, 278, 277, 276, 275, 274, 273, 272,
         287, 286, 285, 284, 283, 282, 281, 280,
         295, 294, 293, 292, 291, 290, 289, 288,
         303, 302, 301, 300, 299, 298, 297, 296,
         311, 310, 309, 308, 307, 306, 305, 304,
         319, 318, 317, 316, 315, 314, 313, 312,
         327, 326, 325, 324, 323, 322, 321, 320,
         335, 334, 333, 332, 331, 330, 329, 328,
         343, 342, 341, 340, 339, 338, 337, 336,
         351, 350, 349, 348, 347, 346, 345, 344,
         359, 358, 357, 356, 355, 354, 353, 352,
         367, 366, 365, 364, 363, 362, 361, 360,
         375, 374, 373, 372, 371, 370, 369, 368,
         383, 382, 381, 380, 379, 378, 377, 376,
         391, 390, 389, 388, 387, 386, 385, 384,
         399, 398, 397, 396, 395, 394, 393, 392,
         407, 406, 405, 404, 403, 402, 401, 400,
         415, 414, 413, 412, 411, 410, 409, 408,
         423, 422, 421, 420, 419, 418, 417, 416,
         431, 430, 429, 428, 427, 426, 425, 424,
         439, 438, 437, 436, 435, 434, 433, 432,
         447, 446, 445, 444, 443, 442, 441, 440,
         455, 454, 453, 452, 451, 450, 449, 448,
         463, 462, 461, 460, 459, 458, 457, 456,
         471, 470, 469, 468, 467, 466, 465, 464,
         479, 478, 477, 476, 475, 474, 473, 472,
         487, 486, 485, 484, 483, 482, 481, 480,
         495, 494, 493, 492, 491, 490, 489, 488,
         503, 502, 501, 500, 499, 498, 497, 496,
         511, 510, 509, 508, 507, 506, 505, 504
         ]
 

# ================== 输出日志 ================== #
 
def print_to_textboxs(*args, end="\n"):
    text = " ".join(str(arg) for arg in args)
    log_text.insert(tk.END, text + end)
    root.update()
    log_text.see(tk.END)
 
 
def print_to_textbox(*args):
    text = " ".join(str(arg) for arg in args)
    log_text.insert(tk.END, text + "\n")
    root.update()
    log_text.see(tk.END)
 

# ================== 表格操作模块 ================== #

def count_worksheets(filename):
    workbook = load_workbook(filename)
    worksheet_count = len(workbook.sheetnames)
    return worksheet_count
 

def get_worksheets(filename):
    workbook = load_workbook(filename)
    worksheets = workbook.sheetnames
    return worksheets
 

def row_to_index(row):
    index = 0
    if isinstance(row, int):
        return row - 1
    if row.isdigit():
        return int(row) - 1
    for char in row:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1
 
 
def insert_signal(arr, signal_name):
    for i in range(len(arr)):
        if signal_name in arr[i]:
            arr.insert(i + 1, 'yes')
            i += 1
    return arr


# ========== 全局变量：Sheet和列名管理 ==========
excel_file_path = None  # 当前选择的Excel文件路径
sheet_headers = {}  # 存储每个Sheet的列名：{sheet_name: [headers]}
sheet_check_vars = {}  # 存储每个Sheet的单选状态：{sheet_name: StringVar}
current_active_sheet = None  # 当前选中的Sheet（用于查看）
current_sheet_var = None  # 当前Sheet选择器的StringVar变量
sheet_column_matches = {}  # 存储每个Sheet的列匹配结果：{sheet_name: {field: column_name}}


# ================== 文件操作模块 ================== #

def clear_file(filename):
    print_to_textbox("清除dbc文件数据")
    with open(filename, 'w') as file:
        file.truncate()
 

def get_input_directory():
    global excel_file_path
    input_directory = filedialog.askopenfilename(title="选择文件输入目录")
    input_directory_button_entry.delete(0, tk.END)
    input_directory_button_entry.insert(0, input_directory)
    
    # 选择Excel文件后，自动设置DBC输出目录为Excel文件所在目录
    if input_directory and os.path.exists(input_directory):
        # 获取Excel文件所在目录
        excel_dir = os.path.dirname(input_directory)
        output_directory_button_entry.delete(0, tk.END)
        output_directory_button_entry.insert(0, excel_dir + "/output2.dbc")
        
        # 加载Sheet列表和列名
        load_sheets_and_columns(input_directory)
 

def get_output_directory():
    # 获取当前Excel文件所在目录作为默认打开位置
    excel_path = input_directory_button_entry.get()
    initial_dir = ""
    initial_file = "output2.dbc"
    
    if excel_path and os.path.exists(excel_path):
        initial_dir = os.path.dirname(excel_path)
    
    # 使用asksaveasfilename让用户可以选择目录并编辑文件名
    output_filepath = filedialog.asksaveasfilename(
        title="选择.DBC文件输出路径",
        initialdir=initial_dir,
        initialfile=initial_file,
        defaultextension=".dbc",
        filetypes=[("DBC文件", "*.dbc"), ("所有文件", "*.*")]
    )
    
    if output_filepath:  # 只有用户选择了文件才更新
        output_directory_button_entry.delete(0, tk.END)
        output_directory_button_entry.insert(0, output_filepath)

# ================== 智能列名匹配模块 ================== #

def get_excel_headers(filepath):
    """
    获取Excel文件的列名列表
    """
    try:
        wb = load_workbook(filepath, read_only=True)
        sheet = wb.active
        
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        
        wb.close()
        return headers
    except Exception as e:
        return []


def column_name_to_index(column_name, headers, used_indices=None, exclude_message_length=False, is_signal_length=False):
    """
    将列名转换为列索引（支持模糊搜索）
    如果输入是列号或列字母，直接转换
    如果输入是列名，在headers中进行模糊匹配
    
    Args:
        column_name: 列名或列号
        headers: 列头列表
        used_indices: 已使用的列索引集合，用于防止重复匹配
        exclude_message_length: 是否排除报文长度相关的列（用于信号长度匹配）
        is_signal_length: 是否是信号长度字段的匹配（用于拦截报文长度特征）
    """
    if not column_name:
        return None
    
    if used_indices is None:
        used_indices = set()
    
    column_name = str(column_name).strip()
    
    # 如果是数字，直接返回
    if column_name.isdigit():
        idx = int(column_name) - 1
        # 检查索引是否在有效范围内
        if idx < 0 or idx >= len(headers):
            return None
        return idx if idx not in used_indices else None
    
    # 如果是纯英文列字母（A-Z, AA-ZZ等），且长度不超过2（避免将BMS/VDC/TBOX等误识别为列字母）
    if column_name.isalpha() and len(column_name) <= 2 and all(c.isascii() for c in column_name):
        index = 0
        for char in column_name.upper():
            index = index * 26 + (ord(char) - ord('A') + 1)
        idx = index - 1
        # 检查索引是否在有效范围内
        if idx < 0 or idx >= len(headers):
            return None
        return idx if idx not in used_indices else None
    
    # 如果是列名，在headers中进行模糊匹配
    if not headers:
        return None
    
    # 清理搜索词：去除首尾空格、换行符等
    search_term = column_name.strip().lower()
    search_term_cleaned = search_term.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    search_term_cleaned = ' '.join(search_term_cleaned.split())  # 合并多个空格
    
    # 为每个候选列计算匹配分数
    best_match_idx = None
    best_score = 0
    
    for idx, header in enumerate(headers):
        
        # 跳过已使用的列
        if idx in used_indices:
            continue
        
        # 如果需要排除报文长度列，检查当前列头是否包含相关关键词
        if exclude_message_length:
            header_lower = header.lower()
            # 排除包含“报文长度”、“Msg Length”、“Message Length”等的列
            if any(kw in header_lower for kw in ['msg length', 'message length', '报文长度']):
                continue  # 跳过该列，不将其作为候选
        
        # 清理列头
        cleaned_header = header.lower().replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        cleaned_header = ' '.join(cleaned_header.split())
        
        # 跳过空列头
        if not cleaned_header:
            continue
        
        score = calculate_match_score(search_term_cleaned, cleaned_header, is_signal_length=is_signal_length)
        
        if score > best_score:
            best_score = score
            best_match_idx = idx
    
    # 只有分数达到阈值才认为匹配成功（提高到50分，避免误匹配）
    if best_score >= 50 and best_match_idx is not None:
        return best_match_idx
    
    return None


def calculate_match_score(search_term, header, is_signal_length=False):
    """
    计算搜索词与列头的匹配分数（0-100）
    
    评分规则：
    - 精确匹配：100分
    - 包含匹配：70-90分
    - 关键词匹配：40-60分
    - 中文模糊匹配：30-50分
    
    Args:
        search_term: 搜索词
        header: 列头
        is_signal_length: 是否是信号长度字段的匹配（用于拦截报文长度特征）
    """
    import re
    
    # 特殊规则：如果搜索词同时包含“报文”和“长度”相关词汇，判定为报文长度列
    # 只有在匹配信号长度时才拦截，报文长度列不应该被拦截
    if is_signal_length:
        search_lower = search_term.lower()
        has_message_keyword = any(kw in search_lower for kw in ['报文', 'message', 'msg'])
        has_length_keyword = any(kw in search_lower for kw in ['长度', 'length'])
        
        if has_message_keyword and has_length_keyword:
            # 这是报文长度列的特征，不应作为信号长度列匹配
            return 0
    
    # 1. 精确匹配（100分）
    if search_term == header:
        return 100
    
    # 2. 包含匹配
    if search_term in header:
        # 搜索词完全包含在列头中
        ratio = len(search_term) / len(header)
        return 70 + int(ratio * 20)  # 70-90分
    
    if header in search_term:
        # 列头完全包含在搜索词中
        ratio = len(header) / len(search_term)
        return 70 + int(ratio * 20)  # 70-90分
    
    # 3. 关键词匹配（分词后匹配）
    search_words = search_term.split()
    header_words = header.split()
    
    # 进一步优化：将中英文混合的词拆分，提高匹配率
    # 例如：'报文id' -> ['报文', 'id']
    def split_mixed_words(word_list):
        result = []
        for word in word_list:
            # 提取所有中文部分
            chinese_parts = re.findall(r'[\u4e00-\u9fff]+', word)
            # 提取所有英文/数字部分
            english_parts = re.findall(r'[a-zA-Z0-9]+', word)
            result.extend(chinese_parts)
            result.extend(english_parts)
        return result if result else word_list
    
    search_words_expanded = split_mixed_words(search_words)
    header_words_expanded = split_mixed_words(header_words)
    
    matched_words = 0
    for sw in search_words_expanded:
        for hw in header_words_expanded:
            if sw.lower() in hw.lower() or hw.lower() in sw.lower():
                matched_words += 1
                break
    
    word_match_ratio = 0
    base_score = 0  # 初始化base_score
    
    if matched_words > 0:
        word_match_ratio = matched_words / len(search_words_expanded)
        # 多词搜索时，如果匹配率达到50%以上，给予基础分数
        if len(search_words_expanded) >= 2 and word_match_ratio >= 0.5:
            base_score = 40 + int(word_match_ratio * 20)  # 40-60分
            # 继续检查是否有中文模糊匹配可以提升分数
        elif len(search_words_expanded) >= 2 and word_match_ratio < 0.5:
            # 匹配率太低，但不直接返回0，而是继续尝试中文匹配
            base_score = 0
        elif len(search_words_expanded) == 1:
            # 单个词的匹配，根据匹配情况给分
            base_score = 40 + int(word_match_ratio * 20)
    
    # 新增：整体包含匹配检查（最高优先级）
    # 如果搜索词的核心部分完全包含在列头中，给予高分
    # 例如：'signal name' 包含在 'norminal signal name 信号简称' 中
    search_core_parts = []
    for word in search_words_expanded:
        # 只考虑长度>=2的英文词或任何中文词
        if len(word) >= 2 or re.search(r'[\u4e00-\u9fff]', word):
            search_core_parts.append(word.lower())
    
    if search_core_parts:
        # 检查是否所有核心部分都出现在列头中
        all_found = all(any(part in hw.lower() for hw in header_words_expanded) for part in search_core_parts)
        if all_found and len(search_core_parts) >= 2:
            # 如果所有核心词都找到了，给予高分
            inclusion_score = 80 + min(20, len(search_core_parts) * 5)  # 80-100分
            base_score = max(base_score, inclusion_score)
    
    # 4. 中文模糊匹配（即使英文匹配率低，也要尝试中文匹配）
    chinese_in_search = re.findall(r'[\u4e00-\u9fff]+', search_term)
    chinese_in_header = re.findall(r'[\u4e00-\u9fff]+', header)
    
    chinese_score = 0
    if chinese_in_search and chinese_in_header:
        max_score = 0
        for search_cn in chinese_in_search:
            for header_cn in chinese_in_header:
                score = calculate_chinese_similarity(search_cn, header_cn)
                max_score = max(max_score, score)
        
        if max_score > 0:
            chinese_score = 30 + int(max_score * 20)  # 30-50分
    
    # 综合评分：取英文匹配和中文匹配的较高者
    final_score = max(base_score, chinese_score)
    
    return final_score


def calculate_chinese_similarity(search_cn, header_cn):
    """
    计算两个中文字符串的相似度（0-1.0）
    """
    # 精确相等
    if search_cn == header_cn:
        return 1.0
    
    # 子串关系（双向检查）
    if search_cn in header_cn or header_cn in search_cn:
        shorter = min(len(search_cn), len(header_cn))
        longer = max(len(search_cn), len(header_cn))
        return shorter / longer
    
    # 前缀匹配
    common_prefix_len = 0
    for i in range(min(len(search_cn), len(header_cn))):
        if search_cn[i] == header_cn[i]:
            common_prefix_len += 1
        else:
            break
    
    if common_prefix_len >= 2:
        return common_prefix_len / max(len(search_cn), len(header_cn))
    
    # 新增：字符级别的部分匹配（对于短字符串更有效）
    # 计算共同字符数
    common_chars = sum(1 for c in search_cn if c in header_cn)
    if common_chars > 0:
        char_ratio = common_chars / max(len(search_cn), len(header_cn))
        # 如果有一定比例的字符相同，给予一定分数
        if char_ratio >= 0.5:  # 至少50%的字符相同
            return char_ratio * 0.8  # 降低权重，避免误匹配
    
    return 0


def column_name_to_index_with_fallback(column_name, headers, field_key, used_indices=None):
    """
    将列名转换为列索引，支持回退匹配
    
    Args:
        column_name: 用户输入的列名
        headers: 列头列表
        field_key: 字段类型（如'message_name', 'precision'等）
        used_indices: 已使用的列索引集合
    
    Returns:
        列索引，如果匹配失败返回None
    """
    if used_indices is None:
        used_indices = set()
    
    # 首先尝试直接匹配
    col_idx = column_name_to_index(column_name, headers, used_indices)
    if col_idx is not None:
        return col_idx
    
    # 如果直接匹配失败，尝试回退匹配
    if field_key and field_key in MATCH_RULES:
        best_fallback_idx = None
        best_fallback_score = 0
        
        # 遍历所有未使用的列
        for idx, header in enumerate(headers):
            if idx in used_indices:
                continue
            
            # 清理列头
            header_cleaned = header.lower().replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
            header_cleaned = ' '.join(header_cleaned.split())
            
            # 跳过空列头
            if not header_cleaned:
                continue
            
            # 对每个关键词计算与当前列头的匹配分数
            for keyword in MATCH_RULES[field_key]:
                score = calculate_match_score(keyword.lower(), header_cleaned, is_signal_length=False)
                
                if score > best_fallback_score:
                    best_fallback_score = score
                    best_fallback_idx = idx
        
        # 只有分数>=50才认为匹配成功
        if best_fallback_idx is not None and best_fallback_score >= 50:
            return best_fallback_idx
    
    return None


# 全局变量：列名匹配规则（用于自动检测和回退匹配）
MATCH_RULES = {
    'message_name': ['报文名称', '报文名', '消息名称', '消息名', 'message name', 'message_name', 'msg name', '帧名称', '帧名', 'Msg Name', 'msg name'],
    'message_id': ['报文id', '消息id', 'message id', 'message_id', 'can id', 'can_id', 'id', '帧id', '帧 ID', '报文标识符', 'Msg ID', 'msg id', 'Message ID', '报文ID'],
    'signal_name': ['信号名称', '信号名', 'signal name', 'signal_name', 'sig name', '参数名称', '参数名', 'Signal Name', 'signal name', 'Norminal Signal name', '信号简称', 'Norminal Signal name 信号简称'],
    'start_bit': ['起始位', '起始bit', 'start bit', 'start_bit', '开始位', '起始位置', 'Start Bit', 'start bit', '起始字节', 'Start Byte', 'Start Bit 起始位'],
    'length': ['Bit Length(Bit)信号长度', 'Bit Length', 'bit length', '信号长度', '信号长度(Bit)', '长度', 'length', '位长', '位数'],
    'message_length': ['报文长度', 'Msg Length', 'Msg Length(Byte)', 'Msg Length(Byte) 报文长度', '报文长度(Byte)', 'Message Length', 'message length'],
    'precision': ['精度', '因子', 'precision', 'factor', '分辨率', 'scale', 'Resolution', 'resolution', 'Precision', '精度(Factor)'],
    'offset': ['偏移', '偏移量', 'offset', '初值', '初始值', 'Offset', 'offset'],
    'comment': ['description', '描述', 'Description', 'Signal Description', 'signal description', '信号描述', 'Description 描述', 'Signal Description 信号描述'],
    'direction': ['TBOX', '收发方向', '方向', 'direction', 'tx/rx', 't/r', '收发', '传输方向'],
    'cycle_time': ['报文周期', '周期', 'cycle time', 'cycle_time', 'period', '报文周期时间', '发送周期', 'Msg Cycle Time', 'msg cycle time', '报文周期时间(ms)'],
    'frame_format': ['Frame Format', 'frame format', '帧格式', '帧类型', 'CAN Type', 'can type', 'Frame Format帧格式', 'frame format帧格式', '帧格式Frame Format', '帧格式frame format', 'Frame Format 帧格式', 'FrameFormat', 'frameformat']
}


def auto_detect_columns_silent(filepath):
    """
    静默自动检测Excel文件中的列名并智能匹配（不输出日志）
    返回匹配结果字典，包含列名和列索引
    """
    try:
        wb = load_workbook(filepath, read_only=True)
        sheet = wb.active
        
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        
        wb.close()
        
        if not headers:
            return None
        
        # 使用全局 MATCH_RULES
        
        matched_columns = {
            'message_name': None,
            'message_id': None,
            'signal_name': None,
            'start_bit': None,
            'length': None,
            'message_length': None,
            'precision': None,
            'offset': None,
            'comment': None,
            'direction': None,
            'cycle_time': None,
            'frame_format': None
        }
        
        # 记录哪些列已经被分配
        used_columns = set()
        
        # 按优先级顺序匹配字段（精确匹配的优先）
        field_priority = [
            'message_name', 'message_id', 'signal_name', 
            'start_bit', 'length', 'message_length', 'precision', 'offset',
            'comment', 'direction', 'cycle_time', 'frame_format'
        ]
        
        # 调试日志
        print(f"\n[自动检测] 开始匹配列名，Excel共有{len(headers)}列")
        
        for field in field_priority:
            keywords = MATCH_RULES[field]
            best_match = None
            best_score = 0
            best_col_idx = -1
            
            print(f"\n[自动检测] 匹配字段: {field}")
            print(f"[自动检测] 关键词列表: {keywords[:3]}...")  # 只显示前3个
            
            for col_idx, header in enumerate(headers):
                # 跳过已使用的列
                if col_idx in used_columns:
                    continue
                    
                if not header:
                    continue
                
                header_lower = header.lower()
                score = 0
                
                # 特殊处理：信号长度(length)字段应该排除报文长度相关的列
                if field == 'length':
                    # 检查列头是否包含报文长度特征
                    has_message_kw = any(kw in header_lower for kw in ['msg', 'message', '报文'])
                    has_length_kw = any(kw in header_lower for kw in ['length', '长度'])
                    if has_message_kw and has_length_kw:
                        print(f"[自动检测]   列{col_idx} '{header}' - 检测到报文长度特征，跳过")
                        continue  # 跳过该列
                
                for keyword in keywords:
                    keyword_lower = keyword.lower()
                    if header_lower == keyword_lower:
                        score = 100
                        print(f"[自动检测]   列{col_idx} '{header}' - 精确匹配 '{keyword}'，得分100")
                        break
                    elif keyword_lower in header_lower:
                        new_score = 50
                        # 计算包含比例，比例越高分数越高
                        ratio = len(keyword_lower) / len(header_lower)
                        new_score = 50 + int(ratio * 20)  # 50-70分
                        if new_score > score:
                            score = new_score
                            print(f"[自动检测]   列{col_idx} '{header}' - 关键词'{keyword}'包含在列头中，得分{score}")
                    elif header_lower in keyword_lower:
                        new_score = 30
                        ratio = len(header_lower) / len(keyword_lower)
                        new_score = 30 + int(ratio * 15)  # 30-45分
                        if new_score > score:
                            score = new_score
                            print(f"[自动检测]   列{col_idx} '{header}' - 列头包含在关键词'{keyword}'中，得分{score}")
                
                if score > best_score:
                    best_score = score
                    best_match = header
                    best_col_idx = col_idx
            
            # 只有分数>=50且列未被使用时才分配（提高阈值）
            if best_score >= 50 and best_col_idx not in used_columns:
                matched_columns[field] = best_match
                used_columns.add(best_col_idx)
                print(f"[自动检测] ✓ {field}: 匹配到 '{best_match}' (列{best_col_idx})，得分{best_score}")
            else:
                print(f"[自动检测] ✗ {field}: 未找到合适匹配（最高得分{best_score}）")
        
        return matched_columns
        
    except Exception as e:
        return None


def load_sheets_and_columns(excel_path):
    """
    加载Excel文件的所有Sheet和列名
    创建Sheet选择器UI
    """
    global excel_file_path, sheet_headers, sheet_check_vars, sheet_column_matches
    
    if not excel_path or not os.path.exists(excel_path):
        return
    
    excel_file_path = excel_path
    sheet_headers = {}
    sheet_check_vars = {}
    sheet_column_matches = {}
    global current_sheet_var
    
    try:
        wb = load_workbook(excel_path, read_only=True)
        
        # 清除旧的Sheet选择器（如果存在）
        if hasattr(load_sheets_and_columns, 'sheet_frame'):
            load_sheets_and_columns.sheet_frame.destroy()
        
        # 创建Sheet选择器框架（在日志区和输入区之间）
        sheet_frame = tk.Frame(root)
        sheet_frame.grid(row=1, column=0, sticky='ew', padx=10, pady=5)
        sheet_frame.grid_columnconfigure(0, weight=1)
        load_sheets_and_columns.sheet_frame = sheet_frame
        
        # 标题
        sheet_title = tk.Label(sheet_frame, text="选择Sheet查看列名匹配结果:", font=('Arial', 10, 'bold'))
        sheet_title.grid(row=0, column=0, sticky='w', pady=5)
        
        # 创建单选按钮容器
        checkbox_frame = tk.Frame(sheet_frame)
        checkbox_frame.grid(row=1, column=0, sticky='ew', pady=3)
        
        # 创建全局的StringVar变量用于Radio Button组
        current_sheet_var = tk.StringVar(value="")
        
        # 读取所有Sheet的列名
        sheet_names = wb.sheetnames
        print_to_textbox(f"\n========== Excel文件信息 ==========")
        print_to_textbox(f"文件路径: {excel_path}")
        print_to_textbox(f"Sheet数量: {len(sheet_names)}")
        
        for idx, sheet_name in enumerate(sheet_names):
            sheet = wb[sheet_name]
            headers = []
            for cell in sheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append("")
            sheet_headers[sheet_name] = headers
            
            # 输出该Sheet的列名信息
            non_empty_headers = [h for h in headers if h]
            print_to_textbox(f"\n[Sheet {idx+1}] {sheet_name} (共{len(headers)}列，有效列{len(non_empty_headers)}个)")
            # 只显示前20个列名，避免日志过长
            display_headers = non_empty_headers[:20]
            print_to_textbox(f"  列名预览: {display_headers}")
            if len(non_empty_headers) > 20:
                print_to_textbox(f"  ... 还有{len(non_empty_headers) - 20}个列未显示")
            
            # 创建单选按钮（Radio Button）- 只能选中一个Sheet来查看其匹配结果
            # 注意：这只是用于查看，不影响实际处理（所有Sheet都会被处理）
            rb = tk.Radiobutton(checkbox_frame, text=sheet_name, variable=current_sheet_var, 
                               value=sheet_name, font=('Arial', 9), 
                               command=lambda name=sheet_name: on_sheet_select(name))
            rb.grid(row=0, column=idx, sticky='w', padx=10)
        
        wb.close()
        
        # 默认选中第一个Sheet
        if sheet_names:
            on_sheet_select(sheet_names[0])
        
        print_to_textbox(f"✓ 已加载Excel文件，共{len(sheet_names)}个Sheet")
        
    except Exception as e:
        print_to_textbox(f"❌ 读取Excel失败: {str(e)}")


def on_sheet_select(sheet_name):
    """
    当用户选择Sheet时调用（Radio Button）
    更新当前活动Sheet和列名下拉框，用于查看该Sheet的匹配结果
    """
    global current_active_sheet
    
    current_active_sheet = sheet_name
    print_to_textbox(f"\n========== 切换到Sheet: {sheet_name} ==========")
    print_to_textbox(f"总列数: {len(sheet_headers[sheet_name])}")
    
    # 显示该Sheet的所有列名（完整列表）
    headers = sheet_headers[sheet_name]
    non_empty = [(i, h) for i, h in enumerate(headers) if h]
    print_to_textbox(f"有效列数: {len(non_empty)}")
    print_to_textbox("列名列表:")
    for idx, header in non_empty:
        print_to_textbox(f"  [{idx}] {header}")
    
    update_column_comboboxes_for_sheet(sheet_name)
    print_to_textbox(f"✓ Sheet切换完成，请查看上方列名列表和下方的匹配结果")


def update_column_comboboxes_for_sheet(sheet_name):
    """
    为指定Sheet更新所有列名下拉框
    """
    global sheet_column_matches
    
    if sheet_name not in sheet_headers:
        return
    
    headers = sheet_headers[sheet_name]
    all_combos = [
        input_message_name_entry, input_message_id_entry,
        input_signal_name_entry, input_start_bit_entry,
        input_length_entry, input_precision_entry,
        input_offset_entry, input_comment_entry,
        input_direction_entry, input_cycle_time_entry,
        input_frame_format_entry, input_message_length_entry
    ]
    
    # 更新所有下拉框的选项（只显示当前Sheet的列名）
    for combo in all_combos:
        combo['values'] = [h for h in headers if h]  # 过滤空列名
    
    # 如果该Sheet已经有保存的匹配结果，恢复显示
    if sheet_name in sheet_column_matches:
        matches = sheet_column_matches[sheet_name]
        
        # 验证每个保存的列名是否在当前Sheet中存在
        def validate_and_set(entry, column_name, field_label):
            """验证列名是否存在于当前Sheet中，存在则设置，否则清空并提示"""
            if column_name and column_name in headers:
                entry.set(column_name)
            else:
                if column_name:  # 只有当之前有值但现在不存在时才提示
                    print_to_textbox(f"⚠ [{sheet_name}] '{field_label}'的列名'{column_name}'在当前Sheet中不存在，已清空")
                entry.set('')  # 列名不存在，清空
        
        validate_and_set(input_message_name_entry, matches.get('message_name'), '报文名称')
        validate_and_set(input_message_id_entry, matches.get('message_id'), '报文ID')
        validate_and_set(input_signal_name_entry, matches.get('signal_name'), '信号名称')
        validate_and_set(input_start_bit_entry, matches.get('start_bit'), '起始位')
        validate_and_set(input_length_entry, matches.get('length'), '信号长度')
        validate_and_set(input_precision_entry, matches.get('precision'), '精度')
        validate_and_set(input_offset_entry, matches.get('offset'), '偏移量')
        validate_and_set(input_comment_entry, matches.get('comment'), '注释')
        validate_and_set(input_direction_entry, matches.get('direction'), '收发方向')
        validate_and_set(input_cycle_time_entry, matches.get('cycle_time'), '报文周期')
        validate_and_set(input_frame_format_entry, matches.get('frame_format'), '帧格式')
        validate_and_set(input_message_length_entry, matches.get('message_length'), '报文长度')
    else:
        # 首次加载，清空下拉框
        for combo in all_combos:
            combo.set('')
        
        # 自动检测该Sheet的列名
        auto_detect_for_sheet(sheet_name)


def auto_detect_for_sheet(sheet_name):
    """
    为指定Sheet自动检测列名
    """
    global sheet_column_matches
    
    if sheet_name not in sheet_headers:
        return
    
    headers = sheet_headers[sheet_name]
    
    print_to_textbox(f"\n[{sheet_name}] ========== 开始自动检测列名 ==========")
    
    # 使用现有的自动检测逻辑
    matched = auto_detect_columns_from_headers(headers)
    
    if matched:
        sheet_column_matches[sheet_name] = matched
        print_to_textbox(f"[{sheet_name}] ✓ 自动检测完成，匹配结果:")
        for field, col_name in matched.items():
            field_names = {
                'message_name': '报文名称',
                'message_id': '报文ID',
                'signal_name': '信号名称',
                'start_bit': '起始位',
                'length': '信号长度',
                'precision': '精度',
                'offset': '偏移量',
                'comment': '注释',
                'direction': '收发方向',
                'cycle_time': '报文周期',
                'frame_format': '帧格式',
                'message_length': '报文长度'
            }
            cn_name = field_names.get(field, field)
            print_to_textbox(f"  {cn_name}: {col_name}")
        
        # 更新下拉框显示
        if matched.get('message_name'):
            input_message_name_entry.set(matched['message_name'])
        if matched.get('message_id'):
            input_message_id_entry.set(matched['message_id'])
        if matched.get('signal_name'):
            input_signal_name_entry.set(matched['signal_name'])
        if matched.get('start_bit'):
            input_start_bit_entry.set(matched['start_bit'])
        if matched.get('length'):
            input_length_entry.set(matched['length'])
        if matched.get('precision'):
            input_precision_entry.set(matched['precision'])
        if matched.get('offset'):
            input_offset_entry.set(matched['offset'])
        if matched.get('comment'):
            input_comment_entry.set(matched['comment'])
        if matched.get('direction'):
            input_direction_entry.set(matched['direction'])
        if matched.get('cycle_time'):
            input_cycle_time_entry.set(matched['cycle_time'])
        if matched.get('frame_format'):
            input_frame_format_entry.set(matched['frame_format'])
        if matched.get('message_length'):
            input_message_length_entry.set(matched['message_length'])
        
        print_to_textbox(f"[{sheet_name}] ✓ 匹配结果已填充到下拉框")
    else:
        print_to_textbox(f"[{sheet_name}] ✗ 自动检测失败，请手动选择列名")


def auto_detect_columns_from_headers(headers):
    """
    根据列头列表自动检测列名（从auto_detect_columns_silent提取的逻辑）
    """
    if not headers:
        return None
    
    matched_columns = {
        'message_name': None,
        'message_id': None,
        'signal_name': None,
        'start_bit': None,
        'length': None,
        'message_length': None,
        'precision': None,
        'offset': None,
        'comment': None,
        'direction': None,
        'cycle_time': None,
        'frame_format': None
    }
    
    used_columns = set()
    
    field_priority = [
        'message_name', 'message_id', 'signal_name', 
        'start_bit', 'length', 'message_length', 'precision', 'offset',
        'comment', 'direction', 'cycle_time', 'frame_format'
    ]
    
    for field in field_priority:
        keywords = MATCH_RULES[field]
        best_match = None
        best_score = 0
        best_col_idx = -1
        
        for col_idx, header in enumerate(headers):
            if col_idx in used_columns or not header:
                continue
            
            header_lower = header.lower()
            score = 0
            
            # 信号长度字段排除报文长度列
            if field == 'length':
                has_message_kw = any(kw in header_lower for kw in ['msg', 'message', '报文'])
                has_length_kw = any(kw in header_lower for kw in ['length', '长度'])
                if has_message_kw and has_length_kw:
                    continue
            
            for keyword in keywords:
                keyword_lower = keyword.lower()
                if header_lower == keyword_lower:
                    score = 100
                    break
                elif keyword_lower in header_lower:
                    ratio = len(keyword_lower) / len(header_lower)
                    score = max(score, 50 + int(ratio * 20))
                elif header_lower in keyword_lower:
                    ratio = len(header_lower) / len(keyword_lower)
                    score = max(score, 30 + int(ratio * 15))
            
            if score > best_score:
                best_score = score
                best_match = header
                best_col_idx = col_idx
        
        if best_score >= 50 and best_col_idx not in used_columns:
            matched_columns[field] = best_match
            used_columns.add(best_col_idx)
    
    return matched_columns


def clear_column_comboboxes():
    """清空所有列名下拉框"""
    all_combos = [
        input_message_name_entry, input_message_id_entry,
        input_signal_name_entry, input_start_bit_entry,
        input_length_entry, input_precision_entry,
        input_offset_entry, input_comment_entry,
        input_direction_entry, input_cycle_time_entry,
        input_frame_format_entry, input_message_length_entry
    ]
    
    for combo in all_combos:
        combo.set('')
        combo['values'] = []


def apply_matched_columns_silent(matched_columns):
    """
    将自动匹配的结果应用到输入框中（使用列名而非列号）
    """
    if matched_columns is None:
        return False
    
    fields_mapping = {
        'message_name': input_message_name_entry,
        'message_id': input_message_id_entry,
        'signal_name': input_signal_name_entry,
        'start_bit': input_start_bit_entry,
        'length': input_length_entry,
        'precision': input_precision_entry,
        'offset': input_offset_entry,
        'comment': input_comment_entry,
        'direction': input_direction_entry,
        'cycle_time': input_cycle_time_entry
    }
    
    all_filled = True
    for field, entry_widget in fields_mapping.items():
        col_name = matched_columns.get(field)
        if col_name:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, col_name)
        else:
            if field != 'comment':
                all_filled = False
    
    return all_filled


def auto_detect_and_fill_silent(filepath):
    """
    静默自动检测并保存结果（不直接操作GUI）
    结果保存在sheet_column_matches中，由update_column_comboboxes_for_sheet负责显示
    """
    global sheet_column_matches
    
    # 如果传入了filepath，先加载Sheet和列名
    if filepath and os.path.exists(filepath):
        load_sheets_and_columns(filepath)
        return
    
    # 对每个Sheet进行自动检测
    for sheet_name, headers in sheet_headers.items():
        matched = auto_detect_columns_from_headers(headers)
        if matched:
            sheet_column_matches[sheet_name] = matched
    
    # 更新当前活动Sheet的显示
    if current_active_sheet and current_active_sheet in sheet_column_matches:
        update_column_comboboxes_for_sheet(current_active_sheet)
    
    return True

# ================== 函数操作模块 ================== #

def on_closing():
    if messagebox.askokcancel("退出程序", "确定要退出吗?"):
        root.destroy()
        sys.exit()
 

def set_lsb_key():
    global storage_mode
    storage_mode = M_LSB
    print_to_textbox("信号存储方式为Motorola最低有效字节（LSB）")

def set_msb_key():
    global storage_mode
    storage_mode = M_MSB
    print_to_textbox("信号存储方式为Motorola最高有效字节（MSB）")

def set_intel_key():
    global storage_mode
    storage_mode = INTEL
    print_to_textbox("信号存储方式为INTEL最高有效字节（INTEL）")
 

def has_chinese_or_empty(text):
    if not text:
        return True
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    result = re.search(pattern, text)
    if result:
        return True
    else:
        return False
 

def sort_sg_lines(sg_lines):
    pattern = r':\s*(\d+)'
    key_func = lambda line: int(re.findall(pattern, line)[0]) if re.findall(pattern, line) else 0
    sorted_sg_lines = sorted(sg_lines, key=key_func)
    return sorted_sg_lines
 

def sort_dbc_lines(input_list, output_list):
    input_list = [line.strip() for line in input_list]
    bo_lines = [line for line in input_list if line.startswith('BO_')]
    sg_lines = [line for line in input_list if line.startswith('SG_')]
    bo_sg_dict = {}
    current_bo = None
    for line in input_list:
        if line.startswith('BO_'):
            current_bo = line
            bo_sg_dict[current_bo] = []
        elif line.startswith('SG_'):
            bo_sg_dict[current_bo].append(line)
    for bo, sg_lines in bo_sg_dict.items():
        sorted_sg_lines = sort_sg_lines(sg_lines)
        bo_sg_dict[bo] = sorted_sg_lines
    for bo in bo_lines:
        output_list.append('\n' + bo)
        output_list.extend([' ' + sg_line for sg_line in bo_sg_dict[bo]])
 

def convert_to_dbc(worksheet_list, wb):
    global storage_mode
    dbc_rows = []
    comment_rows = []
    ba_rows = []  # 存储BA_属性行（包括周期时间）
    bu_nodes = set()  # 收集所有节点名称
    
    # 用于跟踪每个CAN ID的收发关系
    # 格式: {can_id: {'tx_node': node_name, 'rx_nodes': [node1, node2, ...]}}
    can_id_relations = {}
    
    worksheet_names = wb.sheetnames
    worksheet_count = len(worksheet_names)
    has_comment_column = worksheet_list[7] is not None
    
    # 保存用户输入的列名（不是索引）
    direction_col_name = worksheet_list[11] if len(worksheet_list) > 11 else None
    cycle_col_name = worksheet_list[12] if len(worksheet_list) > 12 else None
    frame_format_col_name = worksheet_list[13] if len(worksheet_list) > 13 else None
    
    print_to_textbox(f"[DEBUG] worksheet_list长度: {len(worksheet_list)}, 帧格式列名(索引13): '{frame_format_col_name}'")
    
    print_to_textbox("活动表格数量：", worksheet_count, " ,活动表格名数组：", worksheet_names)
    if has_comment_column:
        print_to_textbox("✓ 已启用信号注释功能")
    if direction_col_name:
        print_to_textbox(f"✓ 已启用收发方向识别功能（列名: '{direction_col_name}'）")
    if cycle_col_name:
        print_to_textbox(f"✓ 已启用报文周期时间功能（列名: '{cycle_col_name}'）")
    if frame_format_col_name:
        print_to_textbox(f"✓ 已启用帧格式识别功能（列名: '{frame_format_col_name}'）")
    else:
        print_to_textbox("✓ 未检测到帧格式列，默认使用标准CAN格式")
 
    for index in range(worksheet_count):
        current_id_error = []
        repetition_id_array = []
        repetition_id = 0
        previous_row = None
        empty_row_count = 0  # 每个sheet开始时重置空行计数器
        
        # 使用sheet名作为节点名（去除空格和特殊字符）
        sheet_name = worksheet_names[index]
        node_name = re.sub(r'[^a-zA-Z0-9_]', '_', sheet_name.replace(' ', '_'))
        # 确保节点名不以数字开头
        if node_name and node_name[0].isdigit():
            node_name = 'N_' + node_name
        if not node_name:
            node_name = 'Node_' + str(index)
        
        bu_nodes.add(node_name)
        
        sheet = wb[worksheet_names[index]]
        
        print_to_textbox(f"\n========== 处理Sheet: {sheet_name} (节点: {node_name})==========")
        print_to_textbox(f"表格中的行数为: {sheet.max_row}")
        
        # 【关键修改】为每个Sheet重新检测所有列的索引
        # 读取当前Sheet的列头
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        
        print_to_textbox(f"[{sheet_name}] Sheet列头数量: {len(headers)}")
        print_to_textbox(f"[{sheet_name}] Sheet列头列表: {headers}")
        
        # 在当前Sheet中查找所有列的索引
        current_msg_name_col = None
        current_msg_id_col = None
        current_sig_name_col = None
        current_start_bit_col = None
        current_length_col = None
        current_message_length_col = None  # 新增：报文长度列
        current_precision_col = None
        current_offset_col = None
        current_comment_col = None
        current_direction_col_index = None
        current_cycle_col_index = None
        current_frame_format_col_index = None
        
        # 基本列配置（索引0-7存储的是列名）
        col_configs = [
            ('报文名称', worksheet_list[0]),
            ('报文ID', worksheet_list[1]),
            ('信号名称', worksheet_list[2]),
            ('起始位', worksheet_list[3]),
            ('长度', worksheet_list[4]),
            ('精度', worksheet_list[5]),
            ('偏移量', worksheet_list[6]),
            ('注释', worksheet_list[7])
        ]
        
        # 检查信号长度列是否误用了报文长度关键词
        signal_length_input = worksheet_list[4]
        if signal_length_input:
            signal_lower = signal_length_input.lower()
            # 如果信号长度列包含“报文”、“message”等关键词，提示用户
            if any(kw in signal_lower for kw in ['报文', 'message', 'msg length']):
                print_to_textbox(f"[{sheet_name}] ⚠ 警告: '长度'列配置为'{signal_length_input}'，这看起来像报文长度列")
                print_to_textbox(f"[{sheet_name}] ℹ 提示: '长度'列应该是信号长度（如'Bit Length'、'信号长度'）")
                print_to_textbox(f"[{sheet_name}] ℹ 如需配置报文长度，请使用单独的'报文长度'输入框")
        
        # 可选列配置：报文长度
        optional_col_configs = [
            ('报文长度', 'message_length')  # 从自动匹配结果中获取
        ]
        
        col_indices = []
        used_indices = set()  # 跟踪已使用的列索引，防止重复匹配
        for col_name, col_value in col_configs:
            if col_value:
                print_to_textbox(f"[{sheet_name}] [DEBUG] 匹配{col_name}: 搜索词='{col_value}'")
                
                # 特殊处理：信号长度列应该排除报文长度相关的列
                exclude_message_length = (col_name == '长度')
                is_signal_length = (col_name == '长度')  # 标记是否为信号长度匹配
                
                col_idx = column_name_to_index(col_value, headers, used_indices, 
                                              exclude_message_length=exclude_message_length,
                                              is_signal_length=is_signal_length)
                
                # 如果用户输入的值匹配失败，尝试使用预定义关键词回退匹配
                if col_idx is None:
                    # 获取该字段的预定义关键词列表
                    field_key = None
                    if col_name == '报文名称':
                        field_key = 'message_name'
                    elif col_name == '报文ID':
                        field_key = 'message_id'
                    elif col_name == '信号名称':
                        field_key = 'signal_name'
                    elif col_name == '起始位':
                        field_key = 'start_bit'
                    elif col_name == '长度':
                        field_key = 'length'
                    elif col_name == '精度':
                        field_key = 'precision'
                    elif col_name == '偏移量':
                        field_key = 'offset'
                    elif col_name == '注释':
                        field_key = 'comment'
                    
                    if field_key and field_key in MATCH_RULES:
                        print_to_textbox(f"[{sheet_name}] ℹ '{col_value}' 匹配失败，尝试使用预定义关键词回退匹配...")
                        # 遍历所有关键词，找到最佳匹配
                        best_fallback_idx = None
                        best_fallback_score = 0
                        
                        # 遍历所有未使用的列
                        for idx, header in enumerate(headers):
                            if idx in used_indices:
                                continue
                            
                            # 清理列头
                            header_cleaned = header.lower().replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
                            header_cleaned = ' '.join(header_cleaned.split())
                            
                            # 跳过空列头
                            if not header_cleaned:
                                continue
                            
                            # 对每个关键词计算与当前列头的匹配分数
                            for keyword in MATCH_RULES[field_key]:
                                score = calculate_match_score(keyword.lower(), header_cleaned, is_signal_length=False)
                                
                                if score > best_fallback_score:
                                    best_fallback_score = score
                                    best_fallback_idx = idx
                        
                        if best_fallback_idx is not None and best_fallback_score >= 50:
                            col_idx = best_fallback_idx
                            actual_header = headers[col_idx] if col_idx < len(headers) else "未知"
                            print_to_textbox(f"[{sheet_name}] ✓ {col_name}: 回退匹配到列'{actual_header}' (索引{col_idx})，得分{best_fallback_score}")
                        else:
                            print_to_textbox(f"[{sheet_name}] ✗ {col_name}: 回退匹配失败，所有关键词都无法匹配")
                
                if col_idx is not None:
                    col_indices.append(col_idx)
                    used_indices.add(col_idx)  # 标记该列为已使用
                    # 显示匹配的列名（Sheet中的实际列名）
                    actual_header = headers[col_idx] if col_idx < len(headers) else "未知"
                    if col_value:  # 只在首次匹配成功时显示
                        print_to_textbox(f"[{sheet_name}] ✓ {col_name}: 搜索'{col_value}' -> 匹配到列'{actual_header}' (索引{col_idx})")
                else:
                    col_indices.append(None)
                    print_to_textbox(f"[{sheet_name}] ⚠ {col_name}: 未找到列 '{col_value}'")
                    # 输出可能相关的列
                    related = [(i, h) for i, h in enumerate(headers) if any(kw.lower() in h.lower() for kw in ['报文', '消息', 'msg', 'name'] if col_name in ['报文名称', '信号名称'])]
                    if related and col_name in ['报文名称', '信号名称']:
                        print_to_textbox(f"[{sheet_name}] [提示] 可能相关的列: {related[:3]}")
            else:
                col_indices.append(None)
                print_to_textbox(f"[{sheet_name}] ⚠ {col_name}: 未配置")
        
        # 解包列索引
        current_msg_name_col, current_msg_id_col, current_sig_name_col, current_start_bit_col, \
        current_length_col, current_precision_col, current_offset_col, current_comment_col = col_indices
        
        # 可选列：收发方向、周期时间、帧格式
        if direction_col_name:
            current_direction_col_index = column_name_to_index(direction_col_name, headers, used_indices)
            if current_direction_col_index is not None:
                used_indices.add(current_direction_col_index)
                actual_header = headers[current_direction_col_index] if current_direction_col_index < len(headers) else "未知"
                print_to_textbox(f"[{sheet_name}] ✓ 收发方向列: 搜索'{direction_col_name}' -> 匹配到列'{actual_header}' (索引{current_direction_col_index})")
            else:
                print_to_textbox(f"[{sheet_name}] ⚠ 未找到收发方向列: '{direction_col_name}'")
        
        if cycle_col_name:
            current_cycle_col_index = column_name_to_index(cycle_col_name, headers, used_indices)
            if current_cycle_col_index is not None:
                used_indices.add(current_cycle_col_index)
                actual_header = headers[current_cycle_col_index] if current_cycle_col_index < len(headers) else "未知"
                print_to_textbox(f"[{sheet_name}] ✓ 报文周期列: 搜索'{cycle_col_name}' -> 匹配到列'{actual_header}' (索引{current_cycle_col_index})")
            else:
                print_to_textbox(f"[{sheet_name}] ⚠ 未找到报文周期列: '{cycle_col_name}'")
        
        if frame_format_col_name:
            print_to_textbox(f"[{sheet_name}] [DEBUG] 尝试匹配帧格式列，搜索词: '{frame_format_col_name}'")
            current_frame_format_col_index = column_name_to_index_with_fallback(frame_format_col_name, headers, 'frame_format', used_indices)
            if current_frame_format_col_index is not None:
                used_indices.add(current_frame_format_col_index)
                actual_header = headers[current_frame_format_col_index] if current_frame_format_col_index < len(headers) else "未知"
                print_to_textbox(f"[{sheet_name}] ✓ 帧格式列: 搜索'{frame_format_col_name}' -> 匹配到列'{actual_header}' (索引{current_frame_format_col_index})")
            else:
                print_to_textbox(f"[{sheet_name}] ⚠ 未找到帧格式列: '{frame_format_col_name}'")
                # 输出所有包含'frame'或'格式'的列头，帮助调试
                matching_headers = [(i, h) for i, h in enumerate(headers) if 'frame' in h.lower() or '格式' in h]
                if matching_headers:
                    print_to_textbox(f"[{sheet_name}] [提示] 找到可能相关的列: {matching_headers}")
        
        # 可选列：报文长度（如果用户没有指定，则尝试自动匹配）
        message_length_col_name = worksheet_list[14] if len(worksheet_list) > 14 and worksheet_list[14] else None
        if not message_length_col_name:
            # 尝试自动匹配报文长度列
            auto_matched = auto_detect_columns_silent(excel_file_path)
            if auto_matched and auto_matched.get('message_length'):
                message_length_col_name = auto_matched['message_length']
                print_to_textbox(f"[{sheet_name}] ℹ 自动检测到报文长度列: '{message_length_col_name}'")
        
        if message_length_col_name:
            print_to_textbox(f"[{sheet_name}] [DEBUG] 尝试匹配报文长度列，搜索词: '{message_length_col_name}'")
            print_to_textbox(f"[{sheet_name}] [DEBUG] 已使用的列索引: {used_indices}")
            current_message_length_col = column_name_to_index_with_fallback(message_length_col_name, headers, 'message_length', used_indices)
            if current_message_length_col is not None:
                used_indices.add(current_message_length_col)
                actual_header = headers[current_message_length_col] if current_message_length_col < len(headers) else "未知"
                print_to_textbox(f"[{sheet_name}] ✓ 报文长度列: 搜索'{message_length_col_name}' -> 匹配到列'{actual_header}' (索引{current_message_length_col})")
            else:
                print_to_textbox(f"[{sheet_name}] ⚠ 未找到报文长度列: '{message_length_col_name}'")
                # 输出所有包含'length'或'长度'的列头，帮助调试
                matching_headers = [(i, h) for i, h in enumerate(headers) if 'length' in h.lower() or '长度' in h]
                if matching_headers:
                    print_to_textbox(f"[{sheet_name}] [提示] 找到包含'length/长度'的列: {matching_headers}")
        
        # 检测当前Sheet是否为CAN FD
        is_can_fd = False
        if current_frame_format_col_index is not None:
            # 读取第一行数据来判断帧格式
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                if row and len(row) > current_frame_format_col_index:
                    format_value = row[current_frame_format_col_index]
                    if format_value is not None:
                        format_str = str(format_value).strip().upper()
                        if 'FD' in format_str or 'CAN_FD' in format_str or 'CANFD' in format_str:
                            is_can_fd = True
                            print_to_textbox(f"  [FD] {sheet_name} 检测到CAN FD格式")
                        else:
                            print_to_textbox(f"  [STD] {sheet_name} 使用标准CAN格式")
                    break
        else:
            print_to_textbox(f"  [STD] {sheet_name} 无帧格式列，使用标准CAN格式")
        
        # 检查基本列是否都存在
        if current_msg_name_col is None or current_msg_id_col is None or current_sig_name_col is None:
            missing_cols = []
            if current_msg_name_col is None:
                missing_cols.append("报文名称")
            if current_msg_id_col is None:
                missing_cols.append("报文ID")
            if current_sig_name_col is None:
                missing_cols.append("信号名称")
            print_to_textbox(f"⚠️  [{sheet_name}] 跳过：缺少必要列 ({', '.join(missing_cols)})")
            continue
        
        for row in sheet.iter_rows(min_row=2, max_row=threshold, values_only=True):
 
            if all(cell is None for cell in row):
                empty_row_count += 1
            else:
                empty_row_count = 0
 
            if empty_row_count >= 5:
                print_to_textbox(f"检测到连续{empty_row_count}个空行，停止处理当前Sheet")
                break
 
            current_row = row[current_msg_id_col]
            if (row[current_msg_name_col] is None and row[current_sig_name_col] is None) or row[current_msg_id_col] is None:
                continue
 
            current_row = str(current_row).replace(" ", "")
            
            # 验证current_row是否是有效的十六进制CAN ID
            try:
                can_id_int = int(current_row, 16)
            except ValueError:
                print_to_textbox(f"  [!] 跳过无效行: CAN ID '{current_row}' 不是有效的十六进制数")
                continue
            
            if current_row != previous_row:
                print_to_textboxs(current_row, end=", ")
                search_value = ' ' + str(can_id_int)
                exists_id = any(search_value in row_data for row_data in dbc_rows)
                
                # 确定报文的收发方向和所属节点
                ecu_str = node_name  # 默认使用当前sheet对应的节点名
                direction = 'T'  # 默认为发送
                
                # 如果配置了收发方向列，读取该列的值（使用当前Sheet的列索引）
                if current_direction_col_index is not None:
                    try:
                        direction_cell = row[current_direction_col_index]
                        
                        if direction_cell is not None:
                            direction_str = str(direction_cell).strip().upper()
                            
                            # 判断是否为T（发送）或R（接收）
                            if 'R' in direction_str and 'T' not in direction_str:
                                direction = 'R'
                                print_to_textbox(f"  [RX] {node_name} 接收 CAN ID 0x{can_id_int:X}")
                            elif 'T' in direction_str:
                                direction = 'T'
                                print_to_textbox(f"  [TX] {node_name} 发送 CAN ID 0x{can_id_int:X}")
                            else:
                                print_to_textbox(f"  [?] 方向列值为'{direction_str}'，使用默认Vector__XXX")
                        else:
                            print_to_textbox(f"  [?] 方向列值为None，使用默认Vector__XXX")
                    except Exception as e:
                        print_to_textbox(f"  [!] 读取方向列异常: {e}")
                
                # 记录CAN ID的收发关系
                if can_id_int not in can_id_relations:
                    can_id_relations[can_id_int] = {'tx_node': None, 'rx_nodes': set()}
                
                if direction == 'T':
                    # 发送节点
                    can_id_relations[can_id_int]['tx_node'] = node_name
                    ecu_str = node_name
                else:
                    # 接收节点
                    can_id_relations[can_id_int]['rx_nodes'].add(node_name)
                    ecu_str = node_name
                
                # 检查是否需要创建新的BO_行
                should_create_bo = False
                
                if exists_id:
                    # CAN ID已存在，检查是否冲突
                    existing_tx_node = can_id_relations[can_id_int]['tx_node']
                    
                    if direction == 'T':
                        # 当前是发送节点
                        if existing_tx_node is None:
                            # 还没有发送节点，可以创建
                            should_create_bo = True
                            print_to_textbox(f"  [+] CAN ID 0x{can_id_int:X} 尚无发送节点，{node_name} 作为发送节点")
                        elif existing_tx_node == node_name:
                            # 同一个节点再次发送 → 真正的重复
                            repetition_id = can_id_int
                            repetition_id_array.append(current_row)
                            print_to_textbox(f"  [!] CAN ID 0x{can_id_int:X} 已由 {node_name} 发送 → 判定为重复")
                        else:
                            # 不同节点都要发送 → 冲突
                            repetition_id = can_id_int
                            repetition_id_array.append(current_row)
                            print_to_textbox(f"  [!] CAN ID 0x{can_id_int:X} 已由 {existing_tx_node} 发送，{node_name} 也要发送 → 冲突，判定为重复")
                    else:
                        # 当前是接收节点，不冲突，只需添加信号
                        repetition_id = can_id_int  # 设置为当前ID，让信号添加到已有报文
                        print_to_textbox(f"  [+] CAN ID 0x{can_id_int:X} 由 {existing_tx_node} 发送，{node_name} 接收 → 补全接收关系")
                else:
                    # CAN ID不存在，需要创建新报文
                    should_create_bo = True
                    repetition_id = 0  # 重置
                
                # 如果需要创建新的BO_行
                if should_create_bo:
                    if row[current_msg_name_col] is None:
                        continue
                    msg_name = str(row[current_msg_name_col])
                    
                    # 读取报文长度（支持CAN FD的8-64字节）
                    msg_length = 8  # 默认标准CAN长度
                    if current_message_length_col is not None:
                        try:
                            length_cell = row[current_message_length_col]
                            if length_cell is not None:
                                # 尝试从单元格读取长度值
                                length_value = int(float(str(length_cell).strip()))
                                # CAN FD最大64字节，最小1字节
                                if 1 <= length_value <= 64:
                                    msg_length = length_value
                                else:
                                    print_to_textbox(f"  [!] 报文长度{length_value}超出范围(1-64)，使用默认值8")
                        except Exception as e:
                            print_to_textbox(f"  [!] 读取报文长度失败: {e}，使用默认值8")
                    
                    # 生成BO_行
                    dbc_rows.append(
                        f"\nBO_ {can_id_int} {msg_name}: {msg_length} {ecu_str}")
                    
                    # 如果配置了帧格式列，根据实际内容设置VFrameFormat属性
                    if current_frame_format_col_index is not None:
                        try:
                            format_cell = row[current_frame_format_col_index]
                            if format_cell is not None:
                                format_str = str(format_cell).strip().upper()
                                # 根据帧格式内容设置对应的枚举值
                                # StandardCAN=0, ExtendedCAN=1, StandardCAN_FD=14, ExtendedCAN_FD=15
                                frame_format_value = 0  # 默认StandardCAN
                                if 'FD' in format_str:
                                    if 'EXTENDED' in format_str or 'EXT' in format_str:
                                        frame_format_value = 15  # ExtendedCAN_FD
                                    else:
                                        frame_format_value = 14  # StandardCAN_FD
                                elif 'EXTENDED' in format_str or 'EXT' in format_str:
                                    frame_format_value = 1  # ExtendedCAN
                                
                                ba_rows.append(f'BA_ "VFrameFormat" BO_ {can_id_int} {frame_format_value};')
                                format_names = {0: 'StandardCAN', 1: 'ExtendedCAN', 14: 'StandardCAN_FD', 15: 'ExtendedCAN_FD'}
                                print_to_textbox(f"  [FORMAT] CAN ID 0x{can_id_int:X} 设置为{format_names.get(frame_format_value, 'Unknown')}格式")
                        except Exception as e:
                            print_to_textbox(f"  [!] 读取帧格式失败: {e}")
                    
                    # 如果配置了周期时间列，读取并生成BA_属性（使用当前Sheet的列索引）
                    if current_cycle_col_index is not None:
                        try:
                            cycle_time_cell = row[current_cycle_col_index]
                            if cycle_time_cell is not None:
                                try:
                                    # 转换为字符串并清理
                                    cycle_str = str(cycle_time_cell).strip()
                                    
                                    # 如果为空或无效，跳过
                                    if not cycle_str or cycle_str.lower() in ['none', 'nan', '']:
                                        print_to_textbox(f"  [周期] ⚠ CAN ID 0x{can_id_int:X} 周期值为空，跳过")
                                    else:
                                        # 提取数字部分（去除单位如ms、s等）
                                        # 匹配数字（包括小数）
                                        match = re.search(r'(\d+\.?\d*)', cycle_str)
                                        if match:
                                            cycle_time_value = int(float(match.group(1)))
                                            if cycle_time_value > 0:
                                                # BA_ "GenMsgCycleTime" BO_ <CAN_ID> <周期时间ms>;
                                                ba_rows.append(f'BA_ "GenMsgCycleTime" BO_ {can_id_int} {cycle_time_value};')
                                                print_to_textbox(f"  [周期] ✓ CAN ID 0x{can_id_int:X} 周期={cycle_time_value}ms (原始值: '{cycle_str}')")
                                            else:
                                                print_to_textbox(f"  [周期] ⚠ CAN ID 0x{can_id_int:X} 周期值<=0，跳过 (原始值: '{cycle_str}')")
                                        else:
                                            print_to_textbox(f"  [周期] ⚠ CAN ID 0x{can_id_int:X} 无法从 '{cycle_str}' 中提取数字，跳过")
                                except Exception as parse_error:
                                    print_to_textbox(f"  [周期] ❌ CAN ID 0x{can_id_int:X} 解析周期值失败: {parse_error} (原始值: '{cycle_time_cell}')")
                            else:
                                print_to_textbox(f"  [周期] ⚠ CAN ID 0x{can_id_int:X} 周期列值为None，跳过")
                        except Exception as e:
                            print_to_textbox(f"  [周期] ❌ 读取周期列异常: {e}")
 
            previous_row = current_row
 
            signal_name_raw = row[current_sig_name_col]
            if signal_name_raw is None or has_chinese_or_empty(str(signal_name_raw)) or 'Reserved' in str(signal_name_raw):
                continue
 
            signal_name = str(signal_name_raw).replace('\n', '')
            signal_name = signal_name.replace(' ', '')
            signal_name = re.sub(r'[^a-zA-Z0-9]', '_', signal_name)
 
            start_bitvalue = str(row[current_start_bit_col])
            try:
                start_bit = int(start_bitvalue)
            except ValueError:
                current_id_error.append(current_row)
                continue
 
            lengthvalue = str(row[current_length_col])
            try:
                length = int(lengthvalue)
            except ValueError:
                current_id_error.append(current_row)
                continue
 
            if storage_mode == M_LSB:
                end_bit = array[array.index(start_bit) - length + 1]
            elif storage_mode == M_MSB:
                end_bit = array[array.index(start_bit) + length - 1]
            else:
                end_bit = start_bit + length - 1
 
            precision = str(row[current_precision_col]) if current_precision_col is not None else '1'
            if precision == 'None':
                precision = '1'
            offset = str(row[current_offset_col]) if current_offset_col is not None else '0'
            if offset == 'None':
                offset = '0'
            
            # 处理注释
            comment_text = ""
            if has_comment_column and current_comment_col is not None and row[current_comment_col] is not None:
                comment_text = str(row[current_comment_col]).strip()
                # 清理注释中的特殊字符，避免DBC格式错误
                comment_text = comment_text.replace('"', "'").replace('\n', ' ').replace('\r', ' ')

            # 确定信号的接收节点列表
            # 核心逻辑：根据方向列的值决定（使用当前Sheet的动态列索引）
            signal_receiver = "Vector__XXX"  # 默认值
            
            # 尝试读取方向列（使用当前Sheet检测到的列索引）
            if current_direction_col_index is not None:
                try:
                    direction_value = row[current_direction_col_index]
                    
                    if direction_value is not None:
                        direction_str = str(direction_value).strip().upper()
                        
                        # R或只包含R（不含T）-> 接收节点
                        if 'R' in direction_str and 'T' not in direction_str:
                            signal_receiver = node_name
                        # T或包含T -> 发送节点
                        elif 'T' in direction_str:
                            signal_receiver = "Vector__XXX"
                except Exception as e:
                    print_to_textbox(f"  [!] 读取信号方向列异常: {e}")
                    signal_receiver = "Vector__XXX"

            if repetition_id == int(current_row, 16):
                exists_nm = any(signal_name in row_data for row_data in dbc_rows)
                if not exists_nm:
                    for i, row_data in enumerate(dbc_rows):
                        if str(repetition_id) in row_data:
                            index_id = i + 1
                            dbc_rows.insert(index_id, f" SG_ {signal_name} : {end_bit}|{length}@0+ ({precision},{offset}) [0|0] \"\" {signal_receiver}")
                            # 添加注释
                            if comment_text:
                                comment_rows.append(f'CM_ SG_ {int(current_row, 16)} {signal_name} "{comment_text}";')
                            break
            else:
                dbc_rows.append(f" SG_ {signal_name} : {end_bit}|{length}@0+ ({precision},{offset}) [0|0] \"\" {signal_receiver}")
                # 添加注释
                if comment_text:
                    comment_rows.append(f'CM_ SG_ {int(current_row, 16)} {signal_name} "{comment_text}";')
 
        print_to_textbox(f'\n{worksheet_names[index]} 转换完成!-----------转换进度{str((index + 1) * 100 / worksheet_count)}%')
        print_to_textbox(f"重复报文 ID : {repetition_id_array}")
        print_to_textbox(f"无法转换 ID : {current_id_error}")
 
    # 生成BU_行，列出所有节点
    bu_line = "BU_: " + " ".join(sorted(bu_nodes))
    
    dbc_rows_output = []
    sort_dbc_lines(dbc_rows, dbc_rows_output)
    print_to_textbox("\n 完成信号排版 -------")
    print_to_textbox(f"共识别 {len(bu_nodes)} 个节点: {', '.join(sorted(bu_nodes))}")
    if ba_rows:
        print_to_textbox(f"✓ 生成 {len(ba_rows)} 条报文周期属性")
    
    # worksheet_list[9] 是输出DBC文件的完整路径
    output_filepath = worksheet_list[9]
    
    # 确保输出路径有效
    if output_filepath is None or output_filepath == '':
        print_to_textbox("❌ 错误: 输出文件路径为空，无法保存DBC文件")
        return
    
    # 使用GBK编码保存文件，支持21003个汉字，被CAN分析工具广泛支持
    # errors='replace' 作为保护机制，将无法编码的字符替换为?
    with open(output_filepath, 'w', encoding='gbk', newline='', errors='replace') as f:
        f.write(title)
        f.write(bu_line + "\n\n")  # 写入节点列表
        
        # 写入报文和信号定义
        f.write('\n'.join(dbc_rows_output) + '\n')
        
        # 写入注释（在报文定义之后）
        if comment_rows:
            f.write('\n'.join(comment_rows) + '\n\n')
            print_to_textbox(f"✓ 已写入 {len(comment_rows)} 条信号注释")
        
        # 写入属性定义和属性值（在注释之后）
        if ba_rows:
            # 第一步：写入所有BA_DEF_属性定义
            f.write('BA_DEF_ BO_  "GenMsgCycleTime" INT 0 65535;\n')
            
            # 如果有CAN FD报文，添加VFrameFormat属性定义
            has_can_fd = any('VFrameFormat' in ba_row for ba_row in ba_rows)
            if has_can_fd:
                f.write('BA_DEF_ BO_  "VFrameFormat" ENUM  "StandardCAN","ExtendedCAN","reserved","J1939PG","StandardCAN_FD","ExtendedCAN_FD";\n')
            
            # 第二步：写入所有BA_DEF_DEF_默认值定义
            f.write('BA_DEF_DEF_  "GenMsgCycleTime" 0;\n')
            if has_can_fd:
                f.write('BA_DEF_DEF_  "VFrameFormat" "StandardCAN";\n')
            
            # 第三步：写入具体的属性值BA_
            f.write('\n'.join(ba_rows) + '\n')
        
        print_to_textbox(f"✓ DBC文件已保存至: {output_filepath}")


def submit_main():
    input_data0 = input_message_name_entry.get().strip()
    input_data1 = input_message_id_entry.get().strip()
    input_data2 = input_signal_name_entry.get().strip()
    input_data3 = input_start_bit_entry.get().strip()
    input_data4 = input_length_entry.get().strip()
    input_data5 = input_precision_entry.get().strip()
    input_data6 = input_offset_entry.get().strip()
    input_data7 = input_comment_entry.get().strip() if input_comment_entry.get().strip() else None
    input_data8 = input_direction_entry.get().strip() if input_direction_entry.get().strip() else None
    input_data9 = input_cycle_time_entry.get().strip() if input_cycle_time_entry.get().strip() else None
    input_data10 = input_frame_format_entry.get().strip() if input_frame_format_entry.get().strip() else None
    input_data11 = input_message_length_entry.get().strip() if input_message_length_entry.get().strip() else None  # 新增：报文长度列
    
    # 清理所有输入中的换行符和多余空格（防止用户粘贴时带入格式）
    def clean_input(text):
        if text is None:
            return None
        # 替换换行符、制表符为空格，然后合并多个空格
        cleaned = text.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
        cleaned = ' '.join(cleaned.split())
        return cleaned if cleaned else None
    
    input_data0 = clean_input(input_data0)
    input_data1 = clean_input(input_data1)
    input_data2 = clean_input(input_data2)
    input_data3 = clean_input(input_data3)
    input_data4 = clean_input(input_data4)
    input_data5 = clean_input(input_data5)
    input_data6 = clean_input(input_data6)
    input_data7 = clean_input(input_data7)
    input_data8 = clean_input(input_data8)
    input_data9 = clean_input(input_data9)
    input_data10 = clean_input(input_data10)
    input_data11 = clean_input(input_data11)  # 清理报文长度列输入
    input_directory = input_directory_button_entry.get()
    output_filepath = output_directory_button_entry.get()
    
    # 如果输出路径为空，使用默认路径
    if not output_filepath or output_filepath.strip() == '':
        excel_dir = os.path.dirname(input_directory) if input_directory and os.path.exists(input_directory) else ""
        if excel_dir:
            output_filepath = os.path.join(excel_dir, "output2.dbc")
        else:
            output_filepath = "output2.dbc"
        print_to_textbox(f"⚠ 输出路径未设置，使用默认路径: {output_filepath}")
 
    worksheet_list = [input_data0, input_data1, input_data2, input_data3, input_data4, input_data5, input_data6, input_data7,
                      input_directory, output_filepath, None, input_data8, input_data9, input_data10, input_data11]
 
    # 获取Excel列头用于列名转换
    headers = get_excel_headers(input_directory)
    
    if not headers:
        print_to_textbox("错误: 无法读取Excel文件列头，请检查文件是否正确")
        return
    
    print_to_textbox(f"Excel列头: {headers}")
    print_to_textbox(f"总列数: {len(headers)}")
    
    # 将列名转换为列索引
    converted_list = []
    field_names = ['报文名称', '报文ID', '信号名称', '起始地址', '信号长度', '精度', '偏移量', '注释']
    field_keys = ['message_name', 'message_id', 'signal_name', 'start_bit', 'length', 'precision', 'offset', 'comment']
    used_indices = set()  # 跟踪已使用的列索引
    
    
    # 处理前8个基本列配置（索引0-7）
    for idx in range(8):
        item = worksheet_list[idx]
        if item is None or item == '':
            converted_list.append(None)
            print_to_textbox(f"⚠ {field_names[idx]}: 未设置")
            continue
        
        # 使用带fallback的匹配函数
        col_index = column_name_to_index_with_fallback(item, headers, field_keys[idx], used_indices)
        
        if col_index is None:
            if idx == 7:  # 注释列是可选的
                print_to_textbox(f"⚠ 注释列: '{item}' 未识别，将不生成注释")
                converted_list.append(None)
                continue
            else:
                print_to_textbox(f"❌ 错误: 无法识别'{field_names[idx]}'的列 '{item}'")
                print_to_textbox(f"   请检查输入是否正确，可用列: {headers}")
                return
        
        if col_index >= len(headers):
            print_to_textbox(f"❌ 错误: '{field_names[idx]}'的列索引 {col_index} 超出范围 (最大索引: {len(headers)-1})")
            print_to_textbox(f"   输入的列: {item}")
            return
        
        converted_list.append(col_index)
        used_indices.add(col_index)  # 标记该列为已使用
        print_to_textbox(f"✓ {field_names[idx]}: '{item}' -> 列索引 {col_index} (列名: {headers[col_index]})")
    
    # 处理方向列（索引11）
    direction_col_name = worksheet_list[11]
    if direction_col_name is None or direction_col_name == '':
        converted_list.append(None)
        print_to_textbox(f"⚠ 收发方向列: 未设置，将默认所有报文为发送(T)")
    else:
        direction_col_index = column_name_to_index_with_fallback(direction_col_name, headers, 'direction', used_indices)
        if direction_col_index is None:
            print_to_textbox(f"⚠ 收发方向列: '{direction_col_name}' 未识别，将默认所有报文为发送(T)")
            converted_list.append(None)
        elif direction_col_index >= len(headers):
            print_to_textbox(f"❌ 错误: 收发方向列索引 {direction_col_index} 超出范围")
            return
        else:
            used_indices.add(direction_col_index)
            converted_list.append(direction_col_index)
            print_to_textbox(f"✓ 收发方向: '{direction_col_name}' -> 列索引 {direction_col_index} (列名: {headers[direction_col_index]})")
    
    # 处理周期列（索引12）
    cycle_col_name = worksheet_list[12]
    if cycle_col_name is None or cycle_col_name == '':
        converted_list.append(None)
        print_to_textbox(f"⚠ 报文周期列: 未设置，将不生成周期属性")
    else:
        cycle_col_index = column_name_to_index_with_fallback(cycle_col_name, headers, 'cycle_time', used_indices)
        if cycle_col_index is None:
            print_to_textbox(f"⚠ 报文周期列: '{cycle_col_name}' 未识别，将不生成周期属性")
            converted_list.append(None)
        elif cycle_col_index >= len(headers):
            print_to_textbox(f"❌ 错误: 报文周期列索引 {cycle_col_index} 超出范围")
            return
        else:
            # 检查是否与方向列冲突
            if len(converted_list) > 8 and converted_list[8] is not None and cycle_col_index == converted_list[8]:
                print_to_textbox(f"⚠ 警告: 报文周期列与收发方向列使用同一列 (索引{cycle_col_index})")
            used_indices.add(cycle_col_index)
            converted_list.append(cycle_col_index)
            print_to_textbox(f"✓ 报文周期: '{cycle_col_name}' -> 列索引 {cycle_col_index} (列名: {headers[cycle_col_index]})")
    
    # 构建最终的worksheet_list
    # 结构：[0-7列配置(列名), 8输入路径, 9输出路径, 10占位符, 11方向列(列名), 12周期列(列名), 13帧格式列(列名), 14报文长度列(列名)]
    worksheet_list_final = [
        worksheet_list[0],  # 0: 报文名称列名（不是索引！）
        worksheet_list[1],  # 1: 报文ID列名（不是索引！）
        worksheet_list[2],  # 2: 信号名称列名（不是索引！）
        worksheet_list[3],  # 3: 起始位列名（不是索引！）
        worksheet_list[4],  # 4: 长度列名（不是索引！）
        worksheet_list[5],  # 5: 精度列名（不是索引！）
        worksheet_list[6],  # 6: 偏移量列名（不是索引！）
        worksheet_list[7],  # 7: 注释列名（不是索引！）
        worksheet_list[8],  # 8: 输入Excel文件路径
        worksheet_list[9],  # 9: 输出DBC文件路径
        None,               # 10: 占位符
        direction_col_name if direction_col_name else None,  # 11: 收发方向列名（不是索引！）
        cycle_col_name if cycle_col_name else None,          # 12: 报文周期列名（不是索引！）
        input_data10 if input_data10 else None,              # 13: 帧格式列名（不是索引！）
        input_data11 if input_data11 else None               # 14: 报文长度列名（不是索引！）
    ]
    
    worksheet_list = worksheet_list_final
    
    print_to_textbox("\n========== 列映射确认 ==========")
    print_to_textbox("输入报文名所在列:", input_data0)
    print_to_textbox("输入报文ID所在列:", input_data1)
    print_to_textbox("输入信号名所在列:", input_data2)
    print_to_textbox("输入起始地址所在列:", input_data3)
    print_to_textbox("输入信号长度所在列:", input_data4)
    print_to_textbox("输入精度所在列:", input_data5)
    print_to_textbox("输入偏移量所在列:", input_data6)
    if worksheet_list[7] is not None:
        print_to_textbox("输入注释所在列:", input_data7)
    else:
        print_to_textbox("注释列: 未设置")
    if worksheet_list[11] is not None:
        print_to_textbox("输入收发方向所在列:", input_data8)
    else:
        print_to_textbox("收发方向列: 未设置（默认所有报文为发送T）")
    if worksheet_list[12] is not None:
        print_to_textbox("输入报文周期所在列:", input_data9)
    else:
        print_to_textbox("报文周期列: 未设置（不生成周期属性）")
    if worksheet_list[13] is not None:
        print_to_textbox("输入帧格式所在列:", input_data10)
    else:
        print_to_textbox("帧格式列: 未设置（默认标准CAN）")
    if worksheet_list[14] is not None:
        print_to_textbox("输入报文长度所在列:", input_data11)
    else:
        print_to_textbox("报文长度列: 未设置（将尝试自动匹配，默认8字节）")
    print_to_textbox("注意：以上列名将在每个Sheet中独立匹配列索引")
    print_to_textbox("================================\n")
    
    print_to_textbox("XLSX文件输入目录:", input_directory)
    print_to_textbox("DBC文件输出路径:", output_filepath)
 
    if storage_mode == M_LSB:
        print_to_textbox("信号存储方式为LSB")
    elif storage_mode == M_MSB:
        print_to_textbox("信号存储方式为MSB")
    else:
        print_to_textbox("信号存储方式为INTEL")
 
    # worksheet_list[8] 是输入Excel文件路径
    wb = openpyxl.load_workbook(worksheet_list[8])
 
    # worksheet_list[9] 是输出DBC文件的完整路径
    filepath = worksheet_list[9]
    
    # 确保filepath不为None
    if filepath is None or filepath == '':
        print_to_textbox("❌ 错误: 输出文件路径为空")
        return
 
    # 使用GBK编码创建初始文件（支持21003个汉字，兼容性好）
    if os.path.exists(filepath):
        with open(filepath, 'w', encoding='gbk', errors='replace') as dbc_file:
            dbc_file.truncate(0)
    else:
        with open(filepath, 'w', encoding='gbk', errors='replace'):
            pass
 
    try:
        convert_to_dbc(worksheet_list, wb)
        print_to_textbox("-----------XlSX转换DBC完毕!")
    except Exception as e:
        error_traceback = traceback.format_exc()
        print_to_textbox("\n！！！代码转化发生错误！！！ \n: ", error_traceback)
        print_to_textbox("\n")
        print_to_textbox("-----------XlSX转换DBC失败!")

root = tk.Tk()
root.geometry("720x620")  # 增加高度以适应Sheet选择器
root.protocol("WM_DELETE_WINDOW", on_closing)
root.title("DBC转化工具 v1.0.0.0")

# 配置grid权重，使窗口可调整大小
root.grid_rowconfigure(0, weight=1)  # 日志区可垂直扩展
root.grid_rowconfigure(1, weight=0)  # Sheet选择器固定高度
root.grid_rowconfigure(2, weight=0)  # 输入区域固定高度
root.grid_columnconfigure(0, weight=1)
 
storage_mode = 1

# ========== 日志区域 ==========
log_frame = tk.Frame(root)
log_frame.grid(row=0, column=0, sticky='nsew', padx=10, pady=5)
log_frame.grid_columnconfigure(0, weight=1)
log_frame.grid_rowconfigure(0, weight=1)

log_text = tk.Text(log_frame, width=72, height=15)
log_text.grid(row=0, column=0, sticky='nsew')

# 添加滚动条
log_scrollbar = tk.Scrollbar(log_frame, command=log_text.yview)
log_scrollbar.grid(row=0, column=1, sticky='ns')
log_text.configure(yscrollcommand=log_scrollbar.set)

print_to_textbox("编译日志 : \r\n默认信号存储方式为最低有效字节（LSB）")
print_to_textbox("提示: 选择Excel文件后将自动识别列名，可直接修改或使用默认值")
print_to_textbox("提示: DBC文件默认名为output2.dbc，可在保存时修改")
print_to_textbox("提示: 可配置收发方向列(T/R)，区分不同CAN路的报文收发方向")
print_to_textbox("提示: 可配置报文周期列，自动生成GenMsgCycleTime属性\n")

# Sheet选择器区域将在load_sheets_and_columns函数中动态创建
# ========== 输入配置区域 ==========
input_frame = tk.Frame(root)
input_frame.grid(row=2, column=0, sticky='nsew', padx=10, pady=5)  # row改为2，为Sheet选择器留出空间
input_frame.grid_columnconfigure(1, weight=1)
input_frame.grid_columnconfigure(3, weight=1)

# 第一行：报文名称、报文ID
input_message_name_label = tk.Label(input_frame, text="输入报文名称所在列:")
input_message_name_label.grid(row=0, column=0, sticky='w', padx=5, pady=3)
input_message_name_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_message_name_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=3)

input_message_id_label = tk.Label(input_frame, text="输入报文ID所在列:")
input_message_id_label.grid(row=0, column=2, sticky='w', padx=5, pady=3)
input_message_id_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_message_id_entry.grid(row=0, column=3, sticky='ew', padx=5, pady=3)

# 第二行：信号名称、起始位
input_signal_name_label = tk.Label(input_frame, text="输入信号名称所在列:")
input_signal_name_label.grid(row=1, column=0, sticky='w', padx=5, pady=3)
input_signal_name_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_signal_name_entry.grid(row=1, column=1, sticky='ew', padx=5, pady=3)

input_start_bit_label = tk.Label(input_frame, text="输入起始地址所在列:")
input_start_bit_label.grid(row=1, column=2, sticky='w', padx=5, pady=3)
input_start_bit_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_start_bit_entry.grid(row=1, column=3, sticky='ew', padx=5, pady=3)

# 第三行：长度、精度
input_length_label = tk.Label(input_frame, text="输入信号长度所在列:")
input_length_label.grid(row=2, column=0, sticky='w', padx=5, pady=3)
input_length_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_length_entry.grid(row=2, column=1, sticky='ew', padx=5, pady=3)

input_precision_label = tk.Label(input_frame, text="输入精度所在列:")
input_precision_label.grid(row=2, column=2, sticky='w', padx=5, pady=3)
input_precision_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_precision_entry.grid(row=2, column=3, sticky='ew', padx=5, pady=3)

# 第四行：偏移量、注释
input_offset_label = tk.Label(input_frame, text="输入偏移量所在列:")
input_offset_label.grid(row=3, column=0, sticky='w', padx=5, pady=3)
input_offset_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_offset_entry.grid(row=3, column=1, sticky='ew', padx=5, pady=3)

input_comment_label = tk.Label(input_frame, text="输入注释所在列:")
input_comment_label.grid(row=3, column=2, sticky='w', padx=5, pady=3)
input_comment_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_comment_entry.grid(row=3, column=3, sticky='ew', padx=5, pady=3)

# 第五行：收发方向、报文周期
input_direction_label = tk.Label(input_frame, text="输入收发方向所在列:")
input_direction_label.grid(row=4, column=0, sticky='w', padx=5, pady=3)
input_direction_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_direction_entry.grid(row=4, column=1, sticky='ew', padx=5, pady=3)

input_cycle_time_label = tk.Label(input_frame, text="输入报文周期所在列:")
input_cycle_time_label.grid(row=4, column=2, sticky='w', padx=5, pady=3)
input_cycle_time_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_cycle_time_entry.grid(row=4, column=3, sticky='ew', padx=5, pady=3)

# 第六行：帧格式、报文长度
input_frame_format_label = tk.Label(input_frame, text="输入帧格式所在列:")
input_frame_format_label.grid(row=5, column=0, sticky='w', padx=5, pady=3)
input_frame_format_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_frame_format_entry.grid(row=5, column=1, sticky='ew', padx=5, pady=3)

input_message_length_label = tk.Label(input_frame, text="输入报文长度所在列:")
input_message_length_label.grid(row=5, column=2, sticky='w', padx=5, pady=3)
input_message_length_entry = ttk.Combobox(input_frame, state='normal', width=25)
input_message_length_entry.grid(row=5, column=3, sticky='ew', padx=5, pady=3)

# 第七行：输入输出路径
input_directory_button_open = tk.Button(input_frame, text="选择XLSX文件输入目录", command=get_input_directory)
input_directory_button_open.grid(row=6, column=0, sticky='w', padx=5, pady=3)
input_directory_button_entry = tk.Entry(input_frame)
input_directory_button_entry.grid(row=6, column=1, columnspan=3, sticky='ew', padx=5, pady=3)

output_directory_button_open = tk.Button(input_frame, text="选择.DBC文件输出路径", command=get_output_directory)
output_directory_button_open.grid(row=7, column=0, sticky='w', padx=5, pady=3)
output_directory_button_entry = tk.Entry(input_frame)
output_directory_button_entry.grid(row=7, column=1, columnspan=3, sticky='ew', padx=5, pady=3)

# 第八行：存储方式和生成按钮
button_frame = tk.Frame(input_frame)
button_frame.grid(row=8, column=0, columnspan=4, sticky='ew', padx=5, pady=10)

button1 = tk.Button(button_frame, text="LSB", command=set_lsb_key, width=10)
button1.pack(side='left', padx=5)

button2 = tk.Button(button_frame, text="MSB", command=set_msb_key, width=10)
button2.pack(side='left', padx=5)

button3 = tk.Button(button_frame, text="INTEL", command=set_intel_key, width=10)
button3.pack(side='left', padx=5)

submit_button = tk.Button(button_frame, text="生成DBC文件", command=submit_main, width=20)
submit_button.pack(side='right', padx=5)
 
root.mainloop()